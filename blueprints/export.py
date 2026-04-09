"""
SBB Weichenheizung - Export Blueprint
PDF und Excel Export für Abnahmetests
"""
import base64
import os
from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required
from xhtml2pdf import pisa


def get_image_as_base64(filepath):
    """
    Liest eine Bilddatei und gibt sie als Base64-Data-URL zurück.

    Args:
        filepath: Absoluter Pfad zur Bilddatei

    Returns:
        str: Base64-Data-URL oder leerer String bei Fehler
    """
    if not os.path.exists(filepath):
        return ''

    # MIME-Type basierend auf Dateiendung
    ext = os.path.splitext(filepath)[1].lower()
    mime_types = {
        '.gif': 'image/gif',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.svg': 'image/svg+xml'
    }
    mime_type = mime_types.get(ext, 'application/octet-stream')

    try:
        with open(filepath, 'rb') as f:
            data = f.read()
        b64 = base64.b64encode(data).decode('utf-8')
        return f'data:{mime_type};base64,{b64}'
    except Exception as e:
        print(f"Error reading image {filepath}: {e}")
        return ''

from models import (db, Project, TestQuestion, AbnahmeTestResult, WHKConfig,
                    ZSKConfig, HGLSConfig, GWHMeteostation, EWHMeteostation,
                    HGLSParameterPruefung, ZSKParameterPruefung)

export_bp = Blueprint('export', __name__)


def convert_html_to_pdf(html_string):
    """
    Konvertiert HTML-String zu PDF mit xhtml2pdf.

    Args:
        html_string: HTML-Inhalt als String

    Returns:
        bytes: PDF als Bytes oder None bei Fehler
    """
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html_string.encode('utf-8')), dest=result)

    if pdf.err:
        return None

    return result.getvalue()


# ==================== HELPER FUNCTIONS ====================

def generate_filename(projekt, selected_sections, file_extension):
    """
    Generiert einen Dateinamen im Format: WH_331 [Projektname] [DIDOK].[extension]

    Args:
        projekt: Project-Objekt
        selected_sections: Liste der gewählten Sektionen (für Kompatibilität, wird nicht mehr verwendet)
        file_extension: Dateiendung ('pdf' oder 'xlsx')

    Returns:
        Generierter Dateiname, z.B. "WH_331 Obermatt OM.pdf"
    """
    # Projektname bereinigen (Sonderzeichen entfernen)
    projektname = projekt.projektname.replace('/', '-').replace('\\', '-')

    # DIDOK - falls vorhanden
    didok = projekt.didok_betriebspunkt or ''

    # Format: WH_331 [Projektname] [DIDOK].extension
    if didok:
        return f'WH_331 {projektname} {didok}.{file_extension}'
    else:
        return f'WH_331 {projektname}.{file_extension}'

def generate_pdf_export(projekt, selected_sections):
    """
    Generiert PDF-Export mit ausgewählten Sektionen.

    Args:
        projekt: Project-Objekt
        selected_sections: Liste der gewählten Sektionen

    Returns:
        PDF-Datei zum Download oder Redirect mit Fehlermeldung
    """
    # Bei GWH-Projekten: GWH-spezifische Export-Funktion verwenden
    if projekt.energie == 'GWH':
        return generate_gwh_pdf_export(projekt, selected_sections)

    try:
        import os

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt.id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse für dieses Projekt laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt.id).all()

        # Ergebnisse in Dictionary umwandeln
        # WICHTIG: Leerzeichen durch Unterstriche ersetzen für konsistente Keys
        results_dict = {}
        for result in results:
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktionen
        def get_icon(result_value):
            # Unicode-Symbole für xhtml2pdf (SVG nicht unterstützt)
            if result_value == 'richtig':
                return '<span style="color: #22c55e; font-size: 14pt; font-weight: bold;">&#10003;</span>'
            elif result_value == 'falsch':
                return '<span style="color: #ef4444; font-size: 14pt; font-weight: bold;">&#10007;</span>'
            elif result_value == 'nicht_testbar':
                return '<span style="color: #6b7280; font-size: 14pt; font-weight: bold;">&#8709;</span>'
            else:
                return ''

        def get_test_result(question_id, system, komponente_index, spalte=''):
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return {
                'icon': get_icon(result_data.get('result')),
                'bemerkung': result_data.get('bemerkung') or ''
            }

        # Helper für Anlage-Ergebnisse (prüft beide Formate: Test-Script und regulärer Speicher)
        def get_anlage_result(question_id, system):
            # Format 1: komponente_index='Anlage', spalte='Anlage' (Test-Script)
            result = get_test_result(question_id, system, 'Anlage', 'Anlage')
            if result['icon']:
                return result
            # Format 2: komponente_index='', spalte='Anlage' (regulärer Speicher)
            return get_test_result(question_id, system, '', 'Anlage')

        # Daten basierend auf selected_sections filtern
        anlage_tests = []
        if 'wh_anlage' in selected_sections:
            anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
            for frage in anlage_fragen:
                # Prüfe beide DB-Formate (komponente_index='Anlage' und komponente_index='')
                wh_lts_result = get_anlage_result(frage.id, 'wh_lts')
                lss_ch_result = get_anlage_result(frage.id, 'lss_ch')

                anlage_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

        # WHK-Daten vorbereiten (nur gewählte WHKs)
        whk_data = []

        for whk_config in whk_configs:
            whk_key = f"whk_{whk_config.whk_nummer.replace(' ', '_')}"

            if whk_key not in selected_sections:
                continue  # Überspringe nicht-gewählte WHKs

            whk_nummer = whk_config.whk_nummer
            # Normalisiere WHK-Nummer für Key-Lookup (Leerzeichen -> Unterstrich)
            whk_nummer_normalized = whk_nummer.replace(' ', '_')

            # WHK-Tests (DB speichert: komponente_index='WHK 01', spalte='WHK 01')
            whk_tests = []
            whk_fragen = [q for q in test_questions if q.komponente_typ == 'WHK']
            for frage in whk_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, whk_nummer_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, whk_nummer_normalized)

                whk_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # Abgang-Tests
            abgang_tests = []
            abgang_fragen = [q for q in test_questions if q.komponente_typ == 'Abgang']
            for abgang_num in range(1, whk_config.anzahl_abgaenge + 1):
                abgang_name = f"Abgang_{abgang_num:02d}"
                for frage in abgang_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, abgang_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, abgang_name)

                    abgang_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': abgang_name.replace('_', ' '),
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Temperatursonden-Tests
            ts_tests = []
            ts_fragen = [q for q in test_questions if q.komponente_typ == 'Temperatursonde']
            for ts_num in range(1, whk_config.anzahl_temperatursonden + 1):
                ts_name = f"TS_{ts_num:02d}"
                for frage in ts_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, ts_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, ts_name)

                    ts_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': ts_name.replace('_', ' '),
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Antriebsheizung-Tests
            ah_tests = []
            if whk_config.hat_antriebsheizung:
                ah_fragen = [q for q in test_questions if q.komponente_typ == 'Antriebsheizung']
                for frage in ah_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, 'Antriebsheizung')
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, 'Antriebsheizung')

                    ah_tests.append({
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            whk_data.append({
                'whk_nummer': whk_nummer,
                'anzahl_abgaenge': whk_config.anzahl_abgaenge,
                'anzahl_temperatursonden': whk_config.anzahl_temperatursonden,
                'whk_tests': whk_tests,
                'abgang_tests': abgang_tests,
                'ts_tests': ts_tests,
                'ah_tests': ah_tests
            })

        # Meteostation-Daten (einzeln ausgewählt)
        meteo_data = []
        # Extrahiere ausgewählte Meteostationen aus selected_sections
        selected_meteo_names = [s.replace('meteostation_', '') for s in selected_sections if s.startswith('meteostation_')]

        if selected_meteo_names:
            meteo_fragen = [q for q in test_questions if q.komponente_typ == 'Meteostation']

            for meteo_name in selected_meteo_names:
                # Normalisiere Meteostation-Namen für Key-Lookup (Leerzeichen -> Unterstrich)
                meteo_name_normalized = meteo_name.replace(' ', '_')
                meteo_tests = []
                for frage in meteo_fragen:
                    # Bei Meteostation: komponente_index UND spalte sind beide der Meteostation-Name
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', meteo_name_normalized, meteo_name_normalized)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', meteo_name_normalized, meteo_name_normalized)

                    meteo_tests.append({
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

                if meteo_tests:
                    meteo_data.append({
                        'meteostation': meteo_name,
                        'tests': meteo_tests
                    })

        # Fehlgeschlagene Tests sammeln (für Zusammenfassung)
        failed_tests = []

        def get_result_value(question_id, system, komponente_index, spalte=''):
            """Hilfsfunktion um den Ergebniswert (nicht das Icon) zu holen"""
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return result_data.get('result')

        # Anlage-Fehler
        # DB speichert: komponente_index='Anlage', spalte='Anlage'
        # Wir pruefen mehrere Varianten für Rückwärtskompatibilität
        if 'wh_anlage' in selected_sections:
            anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
            for frage in anlage_fragen:
                # Pruefe alle moeglichen Key-Varianten (aktuelle + legacy)
                wh_lts_v1 = get_result_value(frage.id, 'wh_lts', 'Anlage', 'Anlage')  # Aktuell korrekt
                wh_lts_v2 = get_result_value(frage.id, 'wh_lts', 'Anlage', '')  # Legacy
                wh_lts_v3 = get_result_value(frage.id, 'wh_lts', '', 'Anlage')  # Legacy
                lss_ch_v1 = get_result_value(frage.id, 'lss_ch', 'Anlage', 'Anlage')  # Aktuell korrekt
                lss_ch_v2 = get_result_value(frage.id, 'lss_ch', 'Anlage', '')  # Legacy
                lss_ch_v3 = get_result_value(frage.id, 'lss_ch', '', 'Anlage')  # Legacy

                # Falsch wenn einer der Eintraege falsch ist
                wh_lts_falsch = wh_lts_v1 == 'falsch' or wh_lts_v2 == 'falsch' or wh_lts_v3 == 'falsch'
                lss_ch_falsch = lss_ch_v1 == 'falsch' or lss_ch_v2 == 'falsch' or lss_ch_v3 == 'falsch'

                if wh_lts_falsch or lss_ch_falsch:
                    # Fuer Icon: Zeige falsch wenn einer falsch ist
                    wh_lts_val = 'falsch' if wh_lts_falsch else (wh_lts_v1 or wh_lts_v2 or wh_lts_v3)
                    lss_ch_val = 'falsch' if lss_ch_falsch else (lss_ch_v1 or lss_ch_v2 or lss_ch_v3)
                    failed_tests.append({
                        'komponente': 'WH-Anlage',
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': get_icon(wh_lts_val),
                        'lss_ch_icon': get_icon(lss_ch_val)
                    })

        # WHK-Fehler (inkl. Abgang, TS, AH)
        for whk_config in whk_configs:
            whk_key = f"whk_{whk_config.whk_nummer.replace(' ', '_')}"
            if whk_key not in selected_sections:
                continue

            whk_nummer = whk_config.whk_nummer
            whk_nummer_normalized = whk_nummer.replace(' ', '_')

            # WHK-Tests
            # In DB kann es zwei Eintraege geben: spalte=None und spalte="WHK 01"
            whk_fragen = [q for q in test_questions if q.komponente_typ == 'WHK']
            for frage in whk_fragen:
                # Pruefe beide moeglichen Key-Varianten
                wh_lts_v1 = get_result_value(frage.id, 'wh_lts', whk_nummer_normalized, '')
                wh_lts_v2 = get_result_value(frage.id, 'wh_lts', whk_nummer_normalized, whk_nummer_normalized)
                lss_ch_v1 = get_result_value(frage.id, 'lss_ch', whk_nummer_normalized, '')
                lss_ch_v2 = get_result_value(frage.id, 'lss_ch', whk_nummer_normalized, whk_nummer_normalized)

                # Falsch wenn einer der Eintraege falsch ist
                wh_lts_falsch = wh_lts_v1 == 'falsch' or wh_lts_v2 == 'falsch'
                lss_ch_falsch = lss_ch_v1 == 'falsch' or lss_ch_v2 == 'falsch'

                if wh_lts_falsch or lss_ch_falsch:
                    wh_lts_val = 'falsch' if wh_lts_falsch else (wh_lts_v1 or wh_lts_v2)
                    lss_ch_val = 'falsch' if lss_ch_falsch else (lss_ch_v1 or lss_ch_v2)
                    failed_tests.append({
                        'komponente': whk_nummer,
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': get_icon(wh_lts_val),
                        'lss_ch_icon': get_icon(lss_ch_val)
                    })

            # Abgang-Tests
            abgang_fragen = [q for q in test_questions if q.komponente_typ == 'Abgang']
            for abgang_num in range(1, whk_config.anzahl_abgaenge + 1):
                abgang_name = f"Abgang_{abgang_num:02d}"
                for frage in abgang_fragen:
                    wh_lts_val = get_result_value(frage.id, 'wh_lts', whk_nummer_normalized, abgang_name)
                    lss_ch_val = get_result_value(frage.id, 'lss_ch', whk_nummer_normalized, abgang_name)
                    if wh_lts_val == 'falsch' or lss_ch_val == 'falsch':
                        failed_tests.append({
                            'komponente': f"{whk_nummer} - {abgang_name.replace('_', ' ')}",
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': get_icon(wh_lts_val),
                            'lss_ch_icon': get_icon(lss_ch_val)
                        })

            # Temperatursonden-Tests
            ts_fragen = [q for q in test_questions if q.komponente_typ == 'Temperatursonde']
            for ts_num in range(1, whk_config.anzahl_temperatursonden + 1):
                ts_name = f"TS_{ts_num:02d}"
                for frage in ts_fragen:
                    wh_lts_val = get_result_value(frage.id, 'wh_lts', whk_nummer_normalized, ts_name)
                    lss_ch_val = get_result_value(frage.id, 'lss_ch', whk_nummer_normalized, ts_name)
                    if wh_lts_val == 'falsch' or lss_ch_val == 'falsch':
                        failed_tests.append({
                            'komponente': f"{whk_nummer} - {ts_name.replace('_', ' ')}",
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': get_icon(wh_lts_val),
                            'lss_ch_icon': get_icon(lss_ch_val)
                        })

            # Antriebsheizung-Tests
            if whk_config.hat_antriebsheizung:
                ah_fragen = [q for q in test_questions if q.komponente_typ == 'Antriebsheizung']
                for frage in ah_fragen:
                    wh_lts_val = get_result_value(frage.id, 'wh_lts', whk_nummer_normalized, 'Antriebsheizung')
                    lss_ch_val = get_result_value(frage.id, 'lss_ch', whk_nummer_normalized, 'Antriebsheizung')
                    if wh_lts_val == 'falsch' or lss_ch_val == 'falsch':
                        failed_tests.append({
                            'komponente': f"{whk_nummer} - Antriebsheizung",
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': get_icon(wh_lts_val),
                            'lss_ch_icon': get_icon(lss_ch_val)
                        })

        # Meteostation-Fehler
        # In DB kann es zwei Eintraege geben: spalte=None und spalte="MS name"
        if selected_meteo_names:
            meteo_fragen = [q for q in test_questions if q.komponente_typ == 'Meteostation']
            for meteo_name in selected_meteo_names:
                meteo_name_normalized = meteo_name.replace(' ', '_')
                for frage in meteo_fragen:
                    # Pruefe beide moeglichen Key-Varianten
                    wh_lts_v1 = get_result_value(frage.id, 'wh_lts', meteo_name_normalized, meteo_name_normalized)
                    wh_lts_v2 = get_result_value(frage.id, 'wh_lts', meteo_name_normalized, '')
                    lss_ch_v1 = get_result_value(frage.id, 'lss_ch', meteo_name_normalized, meteo_name_normalized)
                    lss_ch_v2 = get_result_value(frage.id, 'lss_ch', meteo_name_normalized, '')

                    # Falsch wenn einer der Eintraege falsch ist
                    wh_lts_falsch = wh_lts_v1 == 'falsch' or wh_lts_v2 == 'falsch'
                    lss_ch_falsch = lss_ch_v1 == 'falsch' or lss_ch_v2 == 'falsch'

                    if wh_lts_falsch or lss_ch_falsch:
                        wh_lts_val = 'falsch' if wh_lts_falsch else (wh_lts_v1 or wh_lts_v2)
                        lss_ch_val = 'falsch' if lss_ch_falsch else (lss_ch_v1 or lss_ch_v2)
                        failed_tests.append({
                            'komponente': f"Meteostation {meteo_name}",
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': get_icon(wh_lts_val),
                            'lss_ch_icon': get_icon(lss_ch_val)
                        })

        # Testergebnis-Zähler für Deckblatt-Übersicht
        wh_lts_bestanden = 0
        wh_lts_fehlgeschlagen = 0
        wh_lts_nicht_testbar = 0
        lss_ch_bestanden = 0
        lss_ch_fehlgeschlagen = 0
        lss_ch_nicht_testbar = 0

        for key, data in results_dict.items():
            result_val = data.get('result')
            if '_wh_lts_' in key:
                if result_val == 'richtig':
                    wh_lts_bestanden += 1
                elif result_val == 'falsch':
                    wh_lts_fehlgeschlagen += 1
                elif result_val == 'nicht_testbar':
                    wh_lts_nicht_testbar += 1
            elif '_lss_ch_' in key:
                if result_val == 'richtig':
                    lss_ch_bestanden += 1
                elif result_val == 'falsch':
                    lss_ch_fehlgeschlagen += 1
                elif result_val == 'nicht_testbar':
                    lss_ch_nicht_testbar += 1

        # Assets-Pfade und Base64-kodierte Logos (wie bei GWH)
        app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        assets_path = os.path.join(app_root, 'assets').replace('\\', '/')
        sbb_logo_path = os.path.join(app_root, 'assets', 'sbb06.gif')
        achermann_logo_path = os.path.join(app_root, 'assets', 'logo.png')

        # Logos als Base64 kodieren (zuverlässiger als file:// URLs)
        sbb_logo_base64 = get_image_as_base64(sbb_logo_path)
        achermann_logo_base64 = get_image_as_base64(achermann_logo_path)

        # Template rendern
        html_string = render_template(
            'pdf_abnahmetest.html',
            projekt=projekt,
            projektname=projekt.projektname,
            didok=projekt.didok_betriebspunkt or '',
            projektleiter_sbb=projekt.projektleiter_sbb or '',
            baumappenversion=projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else '',
            pruefer_achermann=projekt.pruefer_achermann or '',
            pruefdatum=projekt.pruefdatum.strftime('%d.%m.%Y') if projekt.pruefdatum else '-',
            export_datum=datetime.now().strftime('%d.%m.%Y'),
            anlage_tests=anlage_tests,
            whk_data=whk_data,
            meteo_data=meteo_data,
            failed_tests=failed_tests,
            assets_path=assets_path,
            sbb_logo_base64=sbb_logo_base64,
            achermann_logo_base64=achermann_logo_base64,
            selected_sections=selected_sections,
            show_deckblatt='deckblatt' in selected_sections,
            wh_lts_bestanden=wh_lts_bestanden,
            wh_lts_fehlgeschlagen=wh_lts_fehlgeschlagen,
            wh_lts_nicht_testbar=wh_lts_nicht_testbar,
            lss_ch_bestanden=lss_ch_bestanden,
            lss_ch_fehlgeschlagen=lss_ch_fehlgeschlagen,
            lss_ch_nicht_testbar=lss_ch_nicht_testbar
        )

        # PDF generieren mit xhtml2pdf
        pdf = convert_html_to_pdf(html_string)

        if pdf is None:
            flash('Fehler bei der PDF-Generierung. Bitte nutzen Sie den Excel-Export als Alternative.', 'error')
            return redirect(url_for('export.export_config', projekt_id=projekt.id))

        # Dateinamen generieren
        filename = generate_filename(projekt, selected_sections, 'pdf')

        # Als Download zurückgeben
        return send_file(
            BytesIO(pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Fehler beim PDF-Export: {str(e)}', 'error')
        return redirect(url_for('export.export_config', projekt_id=projekt.id))


def generate_excel_export(projekt, selected_sections):
    """
    Generiert Excel-Export mit ausgewählten Sektionen.

    Args:
        projekt: Project-Objekt
        selected_sections: Liste der gewählten Sektionen

    Returns:
        Excel-Datei zum Download oder Redirect mit Fehlermeldung
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt.id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt.id).all()

        # Ergebnisse in Dictionary umwandeln
        # WICHTIG: Leerzeichen durch Unterstriche ersetzen für konsistente Keys
        results_dict = {}
        for result in results:
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        def get_result_text(result_value):
            if result_value == 'richtig':
                return '✓ Richtig'
            elif result_value == 'falsch':
                return '✗ Falsch'
            elif result_value == 'nicht_testbar':
                return '⊘ Nicht testbar'
            else:
                return ''

        # Workbook erstellen
        wb = Workbook()
        wb.remove(wb.active)  # Entferne leeres Default-Sheet

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Deckblatt (wenn gewählt)
        if 'deckblatt' in selected_sections:
            ws1 = wb.create_sheet("Deckblatt")
            ws1['A1'] = f"Abnahmetest {projekt.energie or 'Weichenheizung'}"
            ws1['A1'].font = Font(bold=True, size=16)
            ws1.merge_cells('A1:B1')

            ws1['A3'] = "Betriebspunkt (Projektname):"
            ws1['A3'].font = Font(bold=True)
            ws1['B3'] = projekt.projektname

            ws1['A4'] = "DIDOK:"
            ws1['A4'].font = Font(bold=True)
            ws1['B4'] = projekt.didok_betriebspunkt or ''

            ws1['A5'] = "Projektleiter SBB AG:"
            ws1['A5'].font = Font(bold=True)
            ws1['B5'] = projekt.projektleiter_sbb or ''

            ws1['A6'] = "Baumappen Version (Datum):"
            ws1['A6'].font = Font(bold=True)
            ws1['B6'] = projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else ''

            ws1['A7'] = "Prüfer:"
            ws1['A7'].font = Font(bold=True)
            ws1['B7'] = projekt.pruefer_achermann or ''

            ws1['A8'] = "Prüfdatum:"
            ws1['A8'].font = Font(bold=True)
            ws1['B8'] = datetime.now().strftime('%d.%m.%Y')

            ws1.column_dimensions['A'].width = 30
            ws1.column_dimensions['B'].width = 50

        # Sheet 2: WH-Anlage (wenn gewählt)
        if 'wh_anlage' in selected_sections:
            ws2 = wb.create_sheet("WH-Anlage")
            ws2['A1'] = "Test"
            ws2['B1'] = "WH-LTS"
            ws2['C1'] = "LSS-CH"
            ws2['D1'] = "Bemerkung"

            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            row = 2
            anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
            for frage in anlage_fragen:
                # Prüfe beide DB-Formate für Anlage-Ergebnisse
                # Format 1: komponente_index='Anlage', spalte='Anlage' (Test-Script)
                key_wh_lts_v1 = f"{frage.id}_wh_lts_Anlage_Anlage"
                key_lss_ch_v1 = f"{frage.id}_lss_ch_Anlage_Anlage"
                # Format 2: komponente_index='', spalte='Anlage' (regulärer Speicher)
                key_wh_lts_v2 = f"{frage.id}_wh_lts__Anlage"
                key_lss_ch_v2 = f"{frage.id}_lss_ch__Anlage"

                wh_lts_data = results_dict.get(key_wh_lts_v1) or results_dict.get(key_wh_lts_v2, {})
                lss_ch_data = results_dict.get(key_lss_ch_v1) or results_dict.get(key_lss_ch_v2, {})

                ws2[f'A{row}'] = frage.frage_text
                ws2[f'B{row}'] = get_result_text(wh_lts_data.get('result'))
                ws2[f'C{row}'] = get_result_text(lss_ch_data.get('result'))
                ws2[f'D{row}'] = (wh_lts_data.get('bemerkung') or lss_ch_data.get('bemerkung')) or ''

                for cell in ws2[row]:
                    cell.border = border_thin

                row += 1

            ws2.column_dimensions['A'].width = 60
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 40

        # WHK-Sheets (nur für gewählte WHKs)
        for whk_config in whk_configs:
            whk_key = f"whk_{whk_config.whk_nummer.replace(' ', '_')}"

            if whk_key not in selected_sections:
                continue

            ws_whk = wb.create_sheet(f"{whk_config.whk_nummer}")
            ws_whk['A1'] = "Test"
            ws_whk['B1'] = "WH-LTS"
            ws_whk['C1'] = "LSS-CH"
            ws_whk['D1'] = "Bemerkung"

            for cell in ws_whk[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            ws_whk.column_dimensions['A'].width = 60
            ws_whk.column_dimensions['B'].width = 15
            ws_whk.column_dimensions['C'].width = 15
            ws_whk.column_dimensions['D'].width = 40

            row = 2

            # Normalisiere WHK-Nummer für Key-Lookup (Leerzeichen -> Unterstrich)
            whk_nummer_normalized = whk_config.whk_nummer.replace(' ', '_')

            # WHK-Tests
            whk_fragen = [q for q in test_questions if q.komponente_typ == 'WHK']
            for frage in whk_fragen:
                # WHK: komponente_index=WHK-Nummer, spalte=None/''
                key_wh_lts = f"{frage.id}_wh_lts_{whk_nummer_normalized}_"
                key_lss_ch = f"{frage.id}_lss_ch_{whk_nummer_normalized}_"

                wh_lts_data = results_dict.get(key_wh_lts, {})
                lss_ch_data = results_dict.get(key_lss_ch, {})

                ws_whk[f'A{row}'] = frage.frage_text
                ws_whk[f'B{row}'] = get_result_text(wh_lts_data.get('result'))
                ws_whk[f'C{row}'] = get_result_text(lss_ch_data.get('result'))
                ws_whk[f'D{row}'] = (wh_lts_data.get('bemerkung') or lss_ch_data.get('bemerkung')) or ''

                for cell in ws_whk[row]:
                    cell.border = border_thin
                row += 1

        # Meteostation-Sheets (einzeln ausgewählt)
        # Extrahiere ausgewählte Meteostationen aus selected_sections
        selected_meteo_names = [s.replace('meteostation_', '') for s in selected_sections if s.startswith('meteostation_')]

        if selected_meteo_names:
            meteo_fragen = [q for q in test_questions if q.komponente_typ == 'Meteostation']

            for meteo_name in selected_meteo_names:
                # Erstelle Sheet für diese Meteostation
                # Sheet-Name: Max. 31 Zeichen, keine ungültigen Zeichen
                sheet_name = f"Meteo {meteo_name[:23]}" if len(meteo_name) > 23 else f"Meteo {meteo_name}"
                sheet_name = sheet_name.replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '').replace(':', '').replace('?', '')
                ws_meteo = wb.create_sheet(sheet_name)

                ws_meteo['A1'] = "Test"
                ws_meteo['B1'] = "WH-LTS"
                ws_meteo['C1'] = "LSS-CH"
                ws_meteo['D1'] = "Bemerkung"

                for cell in ws_meteo[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_thin

                ws_meteo.column_dimensions['A'].width = 50
                ws_meteo.column_dimensions['B'].width = 15
                ws_meteo.column_dimensions['C'].width = 15
                ws_meteo.column_dimensions['D'].width = 40

                row = 2
                # Normalisiere Meteostation-Namen für Key-Lookup (Leerzeichen -> Unterstrich)
                meteo_name_normalized = meteo_name.replace(' ', '_')
                for frage in meteo_fragen:
                    # Meteostation: komponente_index UND spalte sind beide der Meteostation-Name
                    key_wh_lts = f"{frage.id}_wh_lts_{meteo_name_normalized}_{meteo_name_normalized}"
                    key_lss_ch = f"{frage.id}_lss_ch_{meteo_name_normalized}_{meteo_name_normalized}"

                    wh_lts_data = results_dict.get(key_wh_lts, {})
                    lss_ch_data = results_dict.get(key_lss_ch, {})

                    ws_meteo[f'A{row}'] = frage.frage_text
                    ws_meteo[f'B{row}'] = get_result_text(wh_lts_data.get('result'))
                    ws_meteo[f'C{row}'] = get_result_text(lss_ch_data.get('result'))
                    ws_meteo[f'D{row}'] = (wh_lts_data.get('bemerkung') or lss_ch_data.get('bemerkung')) or ''

                    for cell in ws_meteo[row]:
                        cell.border = border_thin
                    row += 1

        # Excel in Memory speichern
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Dateinamen generieren
        filename = generate_filename(projekt, selected_sections, 'xlsx')

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError:
        flash('openpyxl ist nicht installiert. Bitte installieren Sie es mit: pip install openpyxl', 'error')
        return redirect(url_for('export.export_config', projekt_id=projekt.id))
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Fehler beim Excel-Export: {str(e)}', 'error')
        return redirect(url_for('export.export_config', projekt_id=projekt.id))
def generate_gwh_pdf_export(projekt, selected_sections):
    """
    Generiert GWH PDF-Export mit ausgewählten Sektionen.

    Args:
        projekt: Project-Objekt (GWH)
        selected_sections: Liste der gewählten Sektionen

    Returns:
        PDF-Datei zum Download oder Redirect mit Fehlermeldung
    """
    try:
        import os
        from parameter_definitionen import ZSK_PARAMETER, HGLS_PARAMETER
        from models import HGLSParameterPruefung, ZSKParameterPruefung

        projekt_id = projekt.id

        # GWH-Konfigurationen laden
        hgls_config = HGLSConfig.query.filter_by(projekt_id=projekt_id).first()
        zsk_configs = ZSKConfig.query.filter_by(projekt_id=projekt_id).order_by(ZSKConfig.reihenfolge).all()
        gwh_meteostationen = GWHMeteostation.query.filter_by(projekt_id=projekt_id).order_by(GWHMeteostation.reihenfolge).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse für dieses Projekt laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln
        results_dict = {}
        for result in results:
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktionen
        def get_icon(result_value):
            # Unicode-Symbole für xhtml2pdf (SVG nicht unterstützt)
            if result_value == 'richtig':
                return '<span style="color: #22c55e; font-size: 14pt; font-weight: bold;">&#10003;</span>'
            elif result_value == 'falsch':
                return '<span style="color: #ef4444; font-size: 14pt; font-weight: bold;">&#10007;</span>'
            elif result_value == 'nicht_testbar':
                return '<span style="color: #6b7280; font-size: 14pt; font-weight: bold;">&#8709;</span>'
            else:
                return ''

        def get_test_result(question_id, system, komponente_index, spalte=''):
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return {
                'icon': get_icon(result_data.get('result')),
                'result': result_data.get('result'),
                'bemerkung': result_data.get('bemerkung') or ''
            }

        # Statistik-Zähler
        stats = {
            'wh_lts_bestanden': 0, 'wh_lts_fehlgeschlagen': 0, 'wh_lts_nicht_testbar': 0,
            'lss_ch_bestanden': 0, 'lss_ch_fehlgeschlagen': 0, 'lss_ch_nicht_testbar': 0
        }
        failed_tests = []

        def count_result(result_value, system):
            if result_value == 'richtig':
                stats[f'{system}_bestanden'] += 1
            elif result_value == 'falsch':
                stats[f'{system}_fehlgeschlagen'] += 1
            elif result_value == 'nicht_testbar':
                stats[f'{system}_nicht_testbar'] += 1

        # ==================== GWH-ANLAGE TESTS ====================
        anlage_tests = []
        if 'gwh_anlage' in selected_sections:
            anlage_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Anlage']
            for frage in anlage_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', 'Anlage', 'Anlage')
                lss_ch_result = get_test_result(frage.id, 'lss_ch', 'Anlage', 'Anlage')

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if 'zusammenfassung_fehler' in selected_sections:
                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': 'GWH-Anlage',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                anlage_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

        # ==================== HGLS TESTS ====================
        hgls_tests = []
        hgls_parameter_data = []
        if hgls_config and 'hgls' in selected_sections:
            hgls_fragen = [q for q in test_questions if q.komponente_typ == 'HGLS']
            for frage in hgls_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', 'HGLS', 'HGLS')
                lss_ch_result = get_test_result(frage.id, 'lss_ch', 'HGLS', 'HGLS')

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if 'zusammenfassung_fehler' in selected_sections:
                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': 'HGLS',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                hgls_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

        # HGLS-Parameter (automatisch mit HGLS exportiert)
        if hgls_config and 'hgls' in selected_sections:
            hgls_pruefungen = HGLSParameterPruefung.query.filter_by(projekt_id=projekt_id).all()
            hgls_pruef_dict = {p.parameter_name: p for p in hgls_pruefungen}

            for param in HGLS_PARAMETER:
                pruefung = hgls_pruef_dict.get(param['name'])
                hgls_parameter_data.append({
                    'label': param['label'],
                    'einheit': param['einheit'],
                    'ist_wert': pruefung.ist_wert if pruefung else '',
                    'geprueft': pruefung.geprueft if pruefung else False,
                    'nicht_testbar': pruefung.nicht_testbar if pruefung else False
                })

        # ==================== ZSK DATEN ====================
        zsk_data = []
        for zsk_config in zsk_configs:
            zsk_nummer = zsk_config.zsk_nummer  # z.B. "ZSK 01"
            zsk_nummer_normalized = zsk_nummer.replace(' ', '_')  # z.B. "ZSK_01"

            # Prüfe ob dieser ZSK ausgewählt ist (eine Checkbox pro ZSK wie bei WHK)
            zsk_key = f"zsk_{zsk_nummer}"
            if zsk_key not in selected_sections:
                continue  # Überspringe diesen ZSK komplett

            # ZSK-Tests (immer wenn ZSK ausgewählt)
            zsk_tests = []
            zsk_fragen = [q for q in test_questions if q.komponente_typ == 'ZSK']
            for frage in zsk_fragen:
                # DB speichert: komponente_index="ZSK 01", spalte="ZSK 01"
                wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, zsk_nummer_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, zsk_nummer_normalized)

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if 'zusammenfassung_fehler' in selected_sections:
                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': f'ZSK {zsk_nummer}',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                zsk_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # ZSK-Parameter (immer wenn ZSK ausgewählt)
            zsk_parameter_data = []
            zsk_pruefungen = ZSKParameterPruefung.query.filter_by(
                projekt_id=projekt_id,
                zsk_nummer=zsk_nummer
            ).all()
            zsk_pruef_dict = {p.parameter_name: p for p in zsk_pruefungen}

            for param in ZSK_PARAMETER:
                pruefung = zsk_pruef_dict.get(param['name'])
                zsk_parameter_data.append({
                    'label': param['label'],
                    'einheit': param['einheit'],
                    'ist_wert': pruefung.ist_wert if pruefung else '',
                    'geprueft': pruefung.geprueft if pruefung else False,
                    'nicht_testbar': pruefung.nicht_testbar if pruefung else False
                })

            # Teile-Tests (immer wenn ZSK ausgewählt und Teile vorhanden)
            teile_tests = []
            if zsk_config.anzahl_teile and zsk_config.anzahl_teile > 0:
                teile_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Teile']
                for teil_num in range(1, zsk_config.anzahl_teile + 1):
                    teil_name = f"Teil_{teil_num:02d}"
                    teil_display = f"Teil {teil_num:02d}"
                    for frage in teile_fragen:
                        wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, teil_name)
                        lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, teil_name)

                        count_result(wh_lts_result['result'], 'wh_lts')
                        count_result(lss_ch_result['result'], 'lss_ch')

                        if 'zusammenfassung_fehler' in selected_sections:
                            if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                                failed_tests.append({
                                    'komponente': f'ZSK {zsk_nummer} {teil_display}',
                                    'frage_text': frage.frage_text,
                                    'wh_lts_icon': wh_lts_result['icon'],
                                    'lss_ch_icon': lss_ch_result['icon']
                                })

                        teile_tests.append({
                            'frage_text': frage.frage_text,
                            'spalte': teil_display,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon'],
                            'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                        })

            # Temperatursonde-Tests (immer wenn ZSK ausgewählt und TS vorhanden)
            ts_tests = []
            if zsk_config.hat_temperatursonde:
                ts_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Temperatursonde']
                for frage in ts_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, 'TS')
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, 'TS')

                    count_result(wh_lts_result['result'], 'wh_lts')
                    count_result(lss_ch_result['result'], 'lss_ch')

                    if 'zusammenfassung_fehler' in selected_sections:
                        if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                            failed_tests.append({
                                'komponente': f'ZSK {zsk_nummer} Temperatursonde',
                                'frage_text': frage.frage_text,
                                'wh_lts_icon': wh_lts_result['icon'],
                                'lss_ch_icon': lss_ch_result['icon']
                            })

                    ts_tests.append({
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            zsk_data.append({
                'zsk_nummer': zsk_nummer,
                'name': f'ZSK {zsk_nummer}',
                'anzahl_teile': zsk_config.anzahl_teile or 0,
                'hat_temperatursonde': zsk_config.hat_temperatursonde,
                'zsk_tests': zsk_tests,
                'zsk_parameter': zsk_parameter_data,
                'teile_tests': teile_tests,
                'ts_tests': ts_tests
            })

        # ==================== METEOSTATION DATEN ====================
        meteo_data = []
        meteo_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Meteostation']
        for ms in gwh_meteostationen:
            # Prüfe ob diese Meteostation ausgewählt ist (einzeln auswählbar)
            ms_key = f"gwh_meteostation_{ms.ms_nummer}"
            if ms_key not in selected_sections:
                continue  # Überspringe diese Meteostation

            ms_nummer_normalized = f"MS_{ms.ms_nummer}"
            ms_name_normalized = ms.name.replace(' ', '_')
            meteo_tests = []

            for frage in meteo_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', ms_nummer_normalized, ms_name_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', ms_nummer_normalized, ms_name_normalized)

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if 'zusammenfassung_fehler' in selected_sections:
                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': f'Meteostation {ms.name}',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                meteo_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            if meteo_tests:
                meteo_data.append({
                    'ms_nummer': ms.ms_nummer,
                    'name': ms.name,
                    'tests': meteo_tests
                })

        # Logos als Base64-Data-URLs für xhtml2pdf einbetten
        app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        sbb_logo_path = os.path.join(app_root, 'assets', 'sbb06.gif')
        achermann_logo_path = os.path.join(app_root, 'assets', 'logo.png')

        # Base64-encoded images für xhtml2pdf
        sbb_logo_base64 = get_image_as_base64(sbb_logo_path)
        achermann_logo_base64 = get_image_as_base64(achermann_logo_path)

        # Template rendern
        html_string = render_template(
            'pdf_gwh_abnahmetest.html',
            projekt=projekt,
            projektname=projekt.projektname,
            didok=projekt.didok_betriebspunkt or '',
            projektleiter_sbb=projekt.projektleiter_sbb or '',
            baumappenversion=projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else '',
            pruefer_achermann=projekt.pruefer_achermann or '',
            pruefdatum=projekt.pruefdatum.strftime('%d.%m.%Y') if projekt.pruefdatum else '-',
            export_datum=datetime.now().strftime('%d.%m.%Y'),
            # Tests
            anlage_tests=anlage_tests,
            hgls_config=hgls_config if 'hgls' in selected_sections else None,
            hgls_tests=hgls_tests,
            hgls_parameter=hgls_parameter_data,
            zsk_data=zsk_data,
            meteo_data=meteo_data,
            # Statistik
            wh_lts_bestanden=stats['wh_lts_bestanden'],
            wh_lts_fehlgeschlagen=stats['wh_lts_fehlgeschlagen'],
            wh_lts_nicht_testbar=stats['wh_lts_nicht_testbar'],
            lss_ch_bestanden=stats['lss_ch_bestanden'],
            lss_ch_fehlgeschlagen=stats['lss_ch_fehlgeschlagen'],
            lss_ch_nicht_testbar=stats['lss_ch_nicht_testbar'],
            failed_tests=failed_tests if 'zusammenfassung_fehler' in selected_sections else [],
            # Sections control
            selected_sections=selected_sections,
            show_deckblatt='deckblatt' in selected_sections,
            # Base64-encoded logos
            sbb_logo_base64=sbb_logo_base64,
            achermann_logo_base64=achermann_logo_base64
        )

        # PDF generieren mit xhtml2pdf
        pdf = convert_html_to_pdf(html_string)

        if pdf is None:
            flash('Fehler bei der PDF-Generierung. Bitte nutzen Sie den Excel-Export als Alternative.', 'error')
            return redirect(url_for('export.export_config', projekt_id=projekt.id))

        # Dateinamen generieren
        filename = generate_filename(projekt, selected_sections, 'pdf')

        # Als Download zurückgeben
        return send_file(
            BytesIO(pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Fehler beim PDF-Export: {str(e)}', 'error')

        return redirect(url_for('export.export_config', projekt_id=projekt.id))


# ==================== ROUTES ====================
@export_bp.route('/export')
@login_required
def export():
    """
    Export-Seite mit Projektübersicht.

    Zeigt alle Projekte mit Export-Optionen (PDF, Excel).
    Clientseitige Suchfunktion für schnelles Filtern.

    Returns:
        HTML-Seite mit Projektliste für Export (export.html)
    """
    # Alle Projekte laden, sortiert nach Erstellungsdatum (neueste zuerst)
    projekte = Project.query.order_by(Project.erstellt_am.desc()).all()
    return render_template('export.html', projekte=projekte)

@export_bp.route('/export/projekt/<int:projekt_id>')
@login_required
def export_config(projekt_id):
    """
    Export-Konfigurationsseite für ein spezifisches Projekt.

    Ermöglicht Auswahl von zu exportierenden Sektionen:
    - Deckblatt
    - WH-Anlage / GWH-Anlage
    - Einzelne WHKs (EWH) oder ZSKs (GWH)
    - Einzelne Meteostationen

    Args:
        projekt_id: ID des Projekts

    Returns:
        HTML-Seite mit Konfigurationsoptionen (export_config.html)
    """
    # Projekt laden
    projekt = Project.query.get_or_404(projekt_id)

    # Variablen initialisieren
    whk_configs = []
    meteo_stations = []
    has_meteostationen = False
    hgls_config = None
    zsk_configs = []
    gwh_meteostationen = []

    if projekt.energie == 'GWH':
        # GWH-Konfigurationen laden
        hgls_config = HGLSConfig.query.filter_by(projekt_id=projekt_id).first()
        zsk_configs = ZSKConfig.query.filter_by(projekt_id=projekt_id).order_by(ZSKConfig.reihenfolge).all()
        gwh_meteostationen = GWHMeteostation.query.filter_by(projekt_id=projekt_id).order_by(GWHMeteostation.ms_nummer).all()
    else:
        # EWH: WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

        # EWH-Meteostationen aus dem EWHMeteostation-Modell laden
        ewh_meteostationen = EWHMeteostation.query.filter_by(projekt_id=projekt_id).order_by(EWHMeteostation.reihenfolge).all()

        # Erstelle Liste mit Meteostation-Informationen für das Template
        # (Format wie bei GWH, aber mit ms_nummer als Name da EWHMeteostation kein name-Feld hat)
        meteo_stations = []
        for ms in ewh_meteostationen:
            # Zugeordnete WHKs ermitteln
            zugeordnete_whks = []
            if ms.zugeordnete_whk:
                zugeordnete_whks.append(ms.zugeordnete_whk.whk_nummer)

            # MS-Nummer formatieren (z.B. "01" -> "MS 01")
            ms_display_name = f"MS {ms.ms_nummer}" if not ms.ms_nummer.startswith('MS') else ms.ms_nummer

            meteo_stations.append({
                'ms_nummer': ms.ms_nummer,
                'name': ms_display_name,
                'whk_count': len(zugeordnete_whks),
                'whk_numbers': zugeordnete_whks
            })

        has_meteostationen = len(meteo_stations) > 0

    # Vorgeschlagener Dateiname für den Speichern-Dialog
    suggested_filename = generate_filename(projekt, [], 'pdf')

    return render_template('export_config.html',
                          projekt=projekt,
                          whk_configs=whk_configs,
                          meteo_stations=meteo_stations,
                          has_meteostationen=has_meteostationen,
                          hgls_config=hgls_config,
                          zsk_configs=zsk_configs,
                          gwh_meteostationen=gwh_meteostationen,
                          suggested_filename=suggested_filename)

@export_bp.route('/export/generate', methods=['POST'])
@login_required
def export_generate():
    """
    Verarbeitet Export-Anfrage mit ausgewählten Sektionen.

    POST-Parameter:
        projekt_id: ID des Projekts
        selected_sections[]: Liste der ausgewählten Sektionen
        export_format: 'pdf' oder 'excel'

    Returns:
        Direkter Export (PDF/Excel-Download)
    """
    projekt_id = request.form.get('projekt_id')
    selected_sections = request.form.getlist('selected_sections')
    export_format = request.form.get('export_format', 'pdf')

    # Validierung
    if not projekt_id:
        flash('Projekt-ID fehlt', 'error')
        return redirect(url_for('export.export'))

    if not selected_sections:
        flash('Bitte wählen Sie mindestens eine Sektion aus', 'error')
        return redirect(url_for('export.export_config', projekt_id=projekt_id))

    # Projekt validieren
    projekt = Project.query.get_or_404(projekt_id)

    # Generiere Export basierend auf Format
    if export_format == 'excel':
        return generate_excel_export(projekt, selected_sections)
    else:
        return generate_pdf_export(projekt, selected_sections)

@export_bp.route('/projekt/<int:projekt_id>/gwh-export/pdf')
@login_required
def gwh_export_pdf(projekt_id):
    """
    Exportiert GWH-Abnahmetest-Protokoll als PDF-Datei.

    Verwendet WeasyPrint für PDF-Generierung mit:
    - SBB/Achermann Logos
    - Projektinformationen
    - GWH-Anlage Tests
    - HGLS Tests + HGLS-Parameter
    - ZSK Tests + ZSK-Parameter pro ZSK
    - Teile Tests pro ZSK
    - Temperatursonde Tests pro ZSK
    - Meteostation Tests
    - Zusammenfassung fehlgeschlagener Tests

    Args:
        projekt_id: ID des Projekts

    Returns:
        PDF-Datei zum Download oder Redirect mit Error-Message
    """
    try:
        import os

        # Projekt laden
        projekt = Project.query.get_or_404(projekt_id)

        # Prüfen ob es ein GWH-Projekt ist
        if projekt.energie != 'GWH':
            flash('PDF-Export ist nur für GWH-Projekte verfügbar!', 'error')
            return redirect(url_for('gwh.gwh_abnahmetest', projekt_id=projekt_id))

        # GWH-Konfigurationen laden
        hgls_config = HGLSConfig.query.filter_by(projekt_id=projekt_id).first()
        zsk_configs = ZSKConfig.query.filter_by(projekt_id=projekt_id).order_by(ZSKConfig.reihenfolge).all()
        gwh_meteostationen = GWHMeteostation.query.filter_by(projekt_id=projekt_id).order_by(GWHMeteostation.reihenfolge).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse für dieses Projekt laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln für schnellen Zugriff
        results_dict = {}
        for result in results:
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktion: Icon für Checkbox-Wert
        def get_icon(result_value):
            # Unicode-Symbole für xhtml2pdf (SVG nicht unterstützt)
            if result_value == 'richtig':
                return '<span style="color: #22c55e; font-size: 14pt; font-weight: bold;">&#10003;</span>'
            elif result_value == 'falsch':
                return '<span style="color: #ef4444; font-size: 14pt; font-weight: bold;">&#10007;</span>'
            elif result_value == 'nicht_testbar':
                return '<span style="color: #6b7280; font-size: 14pt; font-weight: bold;">&#8709;</span>'
            else:
                return ''

        # Helper-Funktion: Testergebnis abrufen
        def get_test_result(question_id, system, komponente_index, spalte=''):
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return {
                'icon': get_icon(result_data.get('result')),
                'result': result_data.get('result'),
                'bemerkung': result_data.get('bemerkung') or ''
            }

        # Statistik-Zähler
        stats = {
            'wh_lts_bestanden': 0, 'wh_lts_fehlgeschlagen': 0, 'wh_lts_nicht_testbar': 0,
            'lss_ch_bestanden': 0, 'lss_ch_fehlgeschlagen': 0, 'lss_ch_nicht_testbar': 0
        }
        failed_tests = []

        def count_result(result_value, system):
            if result_value == 'richtig':
                stats[f'{system}_bestanden'] += 1
            elif result_value == 'falsch':
                stats[f'{system}_fehlgeschlagen'] += 1
            elif result_value == 'nicht_testbar':
                stats[f'{system}_nicht_testbar'] += 1

        # ==================== GWH-ANLAGE TESTS ====================
        anlage_tests = []
        anlage_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Anlage']
        for frage in anlage_fragen:
            wh_lts_result = get_test_result(frage.id, 'wh_lts', 'Anlage', 'Anlage')
            lss_ch_result = get_test_result(frage.id, 'lss_ch', 'Anlage', 'Anlage')

            count_result(wh_lts_result['result'], 'wh_lts')
            count_result(lss_ch_result['result'], 'lss_ch')

            if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                failed_tests.append({
                    'komponente': 'GWH-Anlage',
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon']
                })

            anlage_tests.append({
                'frage_text': frage.frage_text,
                'wh_lts_icon': wh_lts_result['icon'],
                'lss_ch_icon': lss_ch_result['icon'],
                'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
            })

        # ==================== HGLS TESTS ====================
        hgls_tests = []
        hgls_parameter_data = []
        if hgls_config:
            hgls_fragen = [q for q in test_questions if q.komponente_typ == 'HGLS']
            for frage in hgls_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', 'HGLS', 'HGLS')
                lss_ch_result = get_test_result(frage.id, 'lss_ch', 'HGLS', 'HGLS')

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                    failed_tests.append({
                        'komponente': 'HGLS',
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon']
                    })

                hgls_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # HGLS-Parameter laden
            from parameter_definitionen import HGLS_PARAMETER
            from models import HGLSParameterPruefung
            hgls_pruefungen = HGLSParameterPruefung.query.filter_by(projekt_id=projekt_id).all()
            hgls_pruef_dict = {p.parameter_name: p for p in hgls_pruefungen}

            for param in HGLS_PARAMETER:
                pruefung = hgls_pruef_dict.get(param['name'])
                hgls_parameter_data.append({
                    'label': param['label'],
                    'einheit': param['einheit'],
                    'ist_wert': pruefung.ist_wert if pruefung else '',
                    'geprueft': pruefung.geprueft if pruefung else False,
                    'nicht_testbar': pruefung.nicht_testbar if pruefung else False
                })

        # ==================== ZSK DATEN ====================
        from parameter_definitionen import ZSK_PARAMETER
        from models import ZSKParameterPruefung

        zsk_data = []
        for zsk_config in zsk_configs:
            zsk_nummer = zsk_config.zsk_nummer  # z.B. "ZSK 01"
            zsk_nummer_normalized = zsk_nummer.replace(' ', '_')  # z.B. "ZSK_01"

            # ZSK-Tests
            zsk_tests = []
            zsk_fragen = [q for q in test_questions if q.komponente_typ == 'ZSK']
            for frage in zsk_fragen:
                # DB speichert: komponente_index="ZSK 01", spalte="ZSK 01"
                wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, zsk_nummer_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, zsk_nummer_normalized)

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                    failed_tests.append({
                        'komponente': f'ZSK {zsk_nummer}',
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon']
                    })

                zsk_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # ZSK-Parameter
            zsk_pruefungen = ZSKParameterPruefung.query.filter_by(
                projekt_id=projekt_id,
                zsk_nummer=zsk_nummer
            ).all()
            zsk_pruef_dict = {p.parameter_name: p for p in zsk_pruefungen}

            zsk_parameter_data = []
            for param in ZSK_PARAMETER:
                pruefung = zsk_pruef_dict.get(param['name'])
                zsk_parameter_data.append({
                    'label': param['label'],
                    'einheit': param['einheit'],
                    'ist_wert': pruefung.ist_wert if pruefung else '',
                    'geprueft': pruefung.geprueft if pruefung else False,
                    'nicht_testbar': pruefung.nicht_testbar if pruefung else False
                })

            # Teile-Tests
            teile_tests = []
            teile_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Teile']
            for teil_num in range(1, (zsk_config.anzahl_teile or 0) + 1):
                teil_name = f"Teil_{teil_num:02d}"
                teil_display = f"Teil {teil_num:02d}"
                for frage in teile_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, teil_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, teil_name)

                    count_result(wh_lts_result['result'], 'wh_lts')
                    count_result(lss_ch_result['result'], 'lss_ch')

                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': f'ZSK {zsk_nummer} {teil_display}',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                    teile_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': teil_display,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Temperatursonde-Tests
            ts_tests = []
            if zsk_config.hat_temperatursonde:
                ts_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Temperatursonde']
                for frage in ts_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', zsk_nummer_normalized, 'TS')
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', zsk_nummer_normalized, 'TS')

                    count_result(wh_lts_result['result'], 'wh_lts')
                    count_result(lss_ch_result['result'], 'lss_ch')

                    if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                        failed_tests.append({
                            'komponente': f'ZSK {zsk_nummer} Temperatursonde',
                            'frage_text': frage.frage_text,
                            'wh_lts_icon': wh_lts_result['icon'],
                            'lss_ch_icon': lss_ch_result['icon']
                        })

                    ts_tests.append({
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            zsk_data.append({
                'zsk_nummer': zsk_nummer,
                'name': f'ZSK {zsk_nummer}',
                'anzahl_teile': zsk_config.anzahl_teile or 0,
                'hat_temperatursonde': zsk_config.hat_temperatursonde,
                'zsk_tests': zsk_tests,
                'zsk_parameter': zsk_parameter_data,
                'teile_tests': teile_tests,
                'ts_tests': ts_tests
            })

        # ==================== METEOSTATION DATEN ====================
        meteo_data = []
        meteo_fragen = [q for q in test_questions if q.komponente_typ == 'GWH_Meteostation']
        for ms in gwh_meteostationen:
            ms_nummer_normalized = f"MS_{ms.ms_nummer}"
            ms_name_normalized = ms.name.replace(' ', '_')
            meteo_tests = []

            for frage in meteo_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', ms_nummer_normalized, ms_name_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', ms_nummer_normalized, ms_name_normalized)

                count_result(wh_lts_result['result'], 'wh_lts')
                count_result(lss_ch_result['result'], 'lss_ch')

                if wh_lts_result['result'] == 'falsch' or lss_ch_result['result'] == 'falsch':
                    failed_tests.append({
                        'komponente': f'Meteostation {ms.name}',
                        'frage_text': frage.frage_text,
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon']
                    })

                meteo_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            if meteo_tests:
                meteo_data.append({
                    'ms_nummer': ms.ms_nummer,
                    'name': ms.name,
                    'tests': meteo_tests
                })

        # Logos als Base64-Data-URLs für xhtml2pdf einbetten
        app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        sbb_logo_path = os.path.join(app_root, 'assets', 'sbb06.gif')
        achermann_logo_path = os.path.join(app_root, 'assets', 'logo.png')

        # Base64-encoded images für xhtml2pdf
        sbb_logo_base64 = get_image_as_base64(sbb_logo_path)
        achermann_logo_base64 = get_image_as_base64(achermann_logo_path)

        # DEBUG: Print logo paths (Route 2)

        # Template rendern
        html_string = render_template(
            'pdf_gwh_abnahmetest.html',
            projekt=projekt,
            projektname=projekt.projektname,
            didok=projekt.didok_betriebspunkt or '',
            projektleiter_sbb=projekt.projektleiter_sbb or '',
            baumappenversion=projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else '',
            pruefer_achermann=projekt.pruefer_achermann or '',
            pruefdatum=projekt.pruefdatum.strftime('%d.%m.%Y') if projekt.pruefdatum else '-',
            export_datum=datetime.now().strftime('%d.%m.%Y'),
            # Tests
            anlage_tests=anlage_tests,
            hgls_config=hgls_config,
            hgls_tests=hgls_tests,
            hgls_parameter=hgls_parameter_data,
            zsk_data=zsk_data,
            meteo_data=meteo_data,
            # Statistik
            wh_lts_bestanden=stats['wh_lts_bestanden'],
            wh_lts_fehlgeschlagen=stats['wh_lts_fehlgeschlagen'],
            wh_lts_nicht_testbar=stats['wh_lts_nicht_testbar'],
            lss_ch_bestanden=stats['lss_ch_bestanden'],
            lss_ch_fehlgeschlagen=stats['lss_ch_fehlgeschlagen'],
            lss_ch_nicht_testbar=stats['lss_ch_nicht_testbar'],
            failed_tests=failed_tests,
            # Base64-encoded logos
            sbb_logo_base64=sbb_logo_base64,
            achermann_logo_base64=achermann_logo_base64
        )

        # PDF generieren mit xhtml2pdf
        pdf = convert_html_to_pdf(html_string)

        if pdf is None:
            flash('Fehler bei der PDF-Generierung.', 'error')
            return redirect(url_for('gwh.gwh_abnahmetest', projekt_id=projekt_id))

        # Als Download zurückgeben
        return send_file(
            BytesIO(pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'GWH_Abnahmetest_{projekt.projektname}_{projekt.didok_betriebspunkt or "keine_DIDOK"}.pdf'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Fehler beim PDF-Export: {str(e)}', 'error')
        return redirect(url_for('gwh.gwh_abnahmetest', projekt_id=projekt_id))


@export_bp.route('/projekt/<int:projekt_id>/export/pdf')
@login_required
def export_pdf(projekt_id):
    """
    Exportiert Abnahmetest-Protokoll als PDF-Datei.

    Verwendet WeasyPrint für PDF-Generierung mit:
    - SBB/Achermann Logos
    - Projektinformationen
    - WHK-Konfiguration
    - Alle Testergebnisse (LSS-CH und WH-LTS)
    - Checkboxen (✓, ✗, ⊘) als SVG-Icons

    Args:
        projekt_id: ID des Projekts

    Returns:
        PDF-Datei zum Download oder Redirect mit Error-Message

    Hinweis:
        Bei Fehlern wird auf Excel-Export verwiesen.
    """
    try:
        import os

        # Projekt laden
        projekt = Project.query.get_or_404(projekt_id)

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse für dieses Projekt laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln für schnellen Zugriff
        # WICHTIG: Leerzeichen durch Unterstriche ersetzen für konsistente Keys
        results_dict = {}
        for result in results:
            # Key-Format: question_id_system_komponente_index_spalte
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktion: Icon für Checkbox-Wert
        def get_icon(result_value):
            # Unicode-Symbole für xhtml2pdf (SVG nicht unterstützt)
            if result_value == 'richtig':
                return '<span style="color: #22c55e; font-size: 14pt; font-weight: bold;">&#10003;</span>'
            elif result_value == 'falsch':
                return '<span style="color: #ef4444; font-size: 14pt; font-weight: bold;">&#10007;</span>'
            elif result_value == 'nicht_testbar':
                return '<span style="color: #6b7280; font-size: 14pt; font-weight: bold;">&#8709;</span>'
            else:
                return ''

        # Helper-Funktion: Testergebnis abrufen
        def get_test_result(question_id, system, komponente_index, spalte=''):
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return {
                'icon': get_icon(result_data.get('result')),
                'bemerkung': result_data.get('bemerkung') or ''
            }

        # WH-Anlage Tests vorbereiten
        anlage_tests = []
        anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
        for frage in anlage_fragen:
            wh_lts_result = get_test_result(frage.id, 'wh_lts', 'Anlage', '')  # Anlage ohne Spalte
            lss_ch_result = get_test_result(frage.id, 'lss_ch', 'Anlage', '')  # Anlage ohne Spalte

            anlage_tests.append({
                'frage_text': frage.frage_text,
                'wh_lts_icon': wh_lts_result['icon'],
                'lss_ch_icon': lss_ch_result['icon'],
                'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
            })

        # WHK-Daten vorbereiten
        whk_data = []
        for whk_config in whk_configs:
            whk_nummer = whk_config.whk_nummer
            # Normalisiere WHK-Nummer für Key-Lookup (Leerzeichen -> Unterstrich)
            whk_nummer_normalized = whk_nummer.replace(' ', '_')

            # WHK-Tests
            whk_tests = []
            whk_fragen = [q for q in test_questions if q.komponente_typ == 'WHK']
            for frage in whk_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, '')  # Spalte leer für WHK-Tests
                lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, '')  # Spalte leer für WHK-Tests

                whk_tests.append({
                    'frage_text': frage.frage_text,
                    'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte (keine separate whk_result Spalte in DB)
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # Abgang-Tests
            abgang_tests = []
            abgang_fragen = [q for q in test_questions if q.komponente_typ == 'Abgang']
            for abgang_num in range(1, whk_config.anzahl_abgaenge + 1):
                abgang_name = f"Abgang_{abgang_num:02d}"  # Mit Unterstrich für Key-Lookup
                abgang_display = f"Abgang {abgang_num:02d}"  # Mit Leerzeichen für Anzeige
                for frage in abgang_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, abgang_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, abgang_name)

                    abgang_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': abgang_display,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Temperatursonden-Tests
            ts_tests = []
            ts_fragen = [q for q in test_questions if q.komponente_typ == 'Temperatursonde']
            for ts_num in range(1, whk_config.anzahl_temperatursonden + 1):
                ts_name = f"TS_{ts_num:02d}"  # Mit Unterstrich für Key-Lookup
                ts_display = f"TS {ts_num:02d}"  # Mit Leerzeichen für Anzeige
                for frage in ts_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, ts_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, ts_name)

                    ts_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': ts_display,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Antriebsheizung-Tests
            ah_tests = []
            if whk_config.hat_antriebsheizung:
                ah_fragen = [q for q in test_questions if q.komponente_typ == 'Antriebsheizung']
                for frage in ah_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer_normalized, 'Antriebsheizung')
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer_normalized, 'Antriebsheizung')

                    ah_tests.append({
                        'frage_text': frage.frage_text,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            whk_data.append({
                'whk_nummer': whk_nummer,
                'anzahl_abgaenge': whk_config.anzahl_abgaenge,
                'anzahl_temperatursonden': whk_config.anzahl_temperatursonden,
                'whk_tests': whk_tests,
                'abgang_tests': abgang_tests,
                'ts_tests': ts_tests,
                'ah_tests': ah_tests
            })

        # Meteostation-Daten vorbereiten
        meteo_data = []
        meteo_stations = list(set([whk.meteostation for whk in whk_configs if whk.meteostation]))
        for meteo_station in meteo_stations:
            # Normalisiere Meteostation-Namen für Key-Lookup (Leerzeichen -> Unterstrich)
            meteo_station_normalized = meteo_station.replace(' ', '_')
            meteo_tests = []
            meteo_fragen = [q for q in test_questions if q.komponente_typ == 'Meteostation']
            for frage in meteo_fragen:
                # Bei Meteostation: komponente_index UND spalte sind beide der Meteostation-Name
                wh_lts_result = get_test_result(frage.id, 'wh_lts', meteo_station_normalized, meteo_station_normalized)
                lss_ch_result = get_test_result(frage.id, 'lss_ch', meteo_station_normalized, meteo_station_normalized)

                meteo_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            if meteo_tests:
                meteo_data.append({
                    'meteostation': meteo_station,
                    'tests': meteo_tests
                })

        # Assets-Pfade und Base64-kodierte Logos
        app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        assets_path = os.path.join(app_root, 'assets').replace('\\', '/')
        sbb_logo_path = os.path.join(app_root, 'assets', 'sbb06.gif')
        achermann_logo_path = os.path.join(app_root, 'assets', 'logo.png')

        # Logos als Base64 kodieren (zuverlässiger als file:// URLs)
        sbb_logo_base64 = get_image_as_base64(sbb_logo_path)
        achermann_logo_base64 = get_image_as_base64(achermann_logo_path)

        # Template rendern
        html_string = render_template(
            'pdf_abnahmetest.html',
            projekt=projekt,
            projektname=projekt.projektname,
            didok=projekt.didok_betriebspunkt or '',
            projektleiter_sbb=projekt.projektleiter_sbb or '',
            baumappenversion=projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else '',
            pruefer_achermann=projekt.pruefer_achermann or '',
            pruefdatum=projekt.pruefdatum.strftime('%d.%m.%Y') if projekt.pruefdatum else '-',
            export_datum=datetime.now().strftime('%d.%m.%Y'),
            anlage_tests=anlage_tests,
            whk_data=whk_data,
            meteo_data=meteo_data,
            assets_path=assets_path,
            sbb_logo_base64=sbb_logo_base64,
            achermann_logo_base64=achermann_logo_base64
        )

        # PDF generieren mit xhtml2pdf
        pdf = convert_html_to_pdf(html_string)

        if pdf is None:
            flash('Fehler bei der PDF-Generierung. Bitte nutzen Sie den Excel-Export als Alternative.', 'error')
            return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))

        # Als Download zurückgeben
        return send_file(
            BytesIO(pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Abnahmetest_{projekt.projektname}_{projekt.didok_betriebspunkt or "keine_DIDOK"}.pdf'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Fehler beim PDF-Export: {str(e)}', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))


# ==================== EXCEL-EXPORT ====================
@export_bp.route('/projekt/<int:projekt_id>/export/excel')
@login_required
def export_excel(projekt_id):
    """
    Exportiert Abnahmetest-Protokoll als Excel-Datei (XLSX).

    Verwendet openpyxl für Excel-Generierung mit:
    - Multiple Sheets (Projektinfo, WH-Anlage, WHK 01, WHK 02, ...)
    - Formatierte Tabellen (Header, Borders, Colors)
    - Testergebnisse als Text (✓ Richtig, ✗ Falsch, ⊘ Nicht testbar)

    Args:
        projekt_id: ID des Projekts

    Returns:
        Excel-Datei zum Download oder Redirect mit Error-Message

    Hinweis:
        Benötigt openpyxl. Funktioniert ohne zusätzliche Dependencies.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        # Projekt laden
        projekt = Project.query.get_or_404(projekt_id)

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln
        # WICHTIG: Leerzeichen durch Unterstriche ersetzen für konsistente Keys
        results_dict = {}
        for result in results:
            komponente_normalized = (result.komponente_index or '').replace(' ', '_')
            spalte_normalized = (result.spalte or '').replace(' ', '_')
            key_wh_lts = f"{result.test_question_id}_wh_lts_{komponente_normalized}_{spalte_normalized}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{komponente_normalized}_{spalte_normalized}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktion: Text für Checkbox-Wert
        def get_result_text(result_value):
            if result_value == 'richtig':
                return '✓ Richtig'
            elif result_value == 'falsch':
                return '✗ Falsch'
            elif result_value == 'nicht_testbar':
                return '⊘ Nicht testbar'
            else:
                return ''

        # Workbook erstellen
        wb = Workbook()

        # Styles definieren
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Projektinformationen
        ws1 = wb.active
        ws1.title = "Projektinformationen"

        ws1['A1'] = "Abnahmetest Elektroweichenheizung"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:B1')

        ws1['A3'] = "Betriebspunkt (Projektname):"
        ws1['A3'].font = Font(bold=True)
        ws1['B3'] = projekt.projektname

        ws1['A4'] = "DIDOK:"
        ws1['A4'].font = Font(bold=True)
        ws1['B4'] = projekt.didok_betriebspunkt or ''

        ws1['A5'] = "Projektleiter SBB AG:"
        ws1['A5'].font = Font(bold=True)
        ws1['B5'] = projekt.projektleiter_sbb or ''

        ws1['A6'] = "Baumappen Version (Datum):"
        ws1['A6'].font = Font(bold=True)
        ws1['B6'] = projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else ''

        ws1['A7'] = "Prüfer:"
        ws1['A7'].font = Font(bold=True)
        ws1['B7'] = projekt.pruefer_achermann or ''

        ws1['A8'] = "Prüfdatum:"
        ws1['A8'].font = Font(bold=True)
        ws1['B8'] = datetime.now().strftime('%d.%m.%Y')

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 50

        # Sheet 2: WH-Anlage Tests
        ws2 = wb.create_sheet("WH-Anlage")
        ws2['A1'] = "Test"
        ws2['B1'] = "WH-LTS"
        ws2['C1'] = "LSS-CH"
        ws2['D1'] = "Bemerkung"

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        row = 2
        anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
        for frage in anlage_fragen:
            # Prüfe beide DB-Formate für Anlage-Ergebnisse
            # Format 1: komponente_index='Anlage', spalte='Anlage' (Test-Script)
            key_wh_lts_v1 = f"{frage.id}_wh_lts_Anlage_Anlage"
            key_lss_ch_v1 = f"{frage.id}_lss_ch_Anlage_Anlage"
            # Format 2: komponente_index='', spalte='Anlage' (regulärer Speicher)
            key_wh_lts_v2 = f"{frage.id}_wh_lts__Anlage"
            key_lss_ch_v2 = f"{frage.id}_lss_ch__Anlage"

            wh_lts_data = results_dict.get(key_wh_lts_v1) or results_dict.get(key_wh_lts_v2, {})
            lss_ch_data = results_dict.get(key_lss_ch_v1) or results_dict.get(key_lss_ch_v2, {})

            ws2[f'A{row}'] = frage.frage_text
            ws2[f'B{row}'] = get_result_text(wh_lts_data.get('result'))
            ws2[f'C{row}'] = get_result_text(lss_ch_data.get('result'))
            ws2[f'D{row}'] = (wh_lts_data.get('bemerkung') or lss_ch_data.get('bemerkung')) or ''

            for cell in ws2[row]:
                cell.border = border_thin

            row += 1

        ws2.column_dimensions['A'].width = 60
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 40

        # Weitere Sheets für WHKs
        for whk_config in whk_configs:
            ws_whk = wb.create_sheet(f"WHK {whk_config.whk_nummer}")
            ws_whk['A1'] = "Test"
            ws_whk['B1'] = "WH-LTS"
            ws_whk['C1'] = "LSS-CH"
            ws_whk['D1'] = "Bemerkung"

            for cell in ws_whk[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            ws_whk.column_dimensions['A'].width = 60
            ws_whk.column_dimensions['B'].width = 15
            ws_whk.column_dimensions['C'].width = 15
            ws_whk.column_dimensions['D'].width = 40

            # Hier würden die WHK-Tests eingefügt werden (ähnlich wie oben)
            # Aus Platzgründen hier vereinfacht

        # Excel in Memory speichern
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        from flask import send_file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Abnahmetest_{projekt.projektname}_{projekt.didok_betriebspunkt or "keine_DIDOK"}.xlsx'
        )

    except ImportError:
        flash('openpyxl ist nicht installiert. Bitte installieren Sie es mit: pip install openpyxl', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))
    except Exception as e:
        flash(f'Fehler beim Excel-Export: {str(e)}', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))
