"""
Script zum Analysieren und Importieren der Testfragen aus dem Excel-File
"""
import sys
import os
import io

# UTF-8 Encoding für Windows-Konsole
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Projekt-Root zum Python-Path hinzufügen
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from app import app, db
from models import TestQuestion

def analyze_excel():
    """Analysiert das Excel-File und zeigt die Struktur"""
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              'Testfragen_Weichenheizung_Komplett.xlsx')

    print(f"Analysiere Excel-File: {excel_path}")
    print("=" * 80)

    # Excel-File laden
    wb = load_workbook(excel_path, data_only=True)

    print(f"\nGefundene Sheets: {wb.sheetnames}")
    print("=" * 80)

    # Jedes Sheet analysieren
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n### Sheet: {sheet_name} ###")
        print(f"Dimensionen: {ws.dimensions}")
        print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")

        # Header-Zeile anzeigen
        if ws.max_row > 0:
            print("\nHeader (erste Zeile):")
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(cell_value)
                print(f"  Spalte {col}: {cell_value}")

            # Erste 3 Datenzeilen als Beispiel
            print("\nErste 3 Datenzeilen:")
            for row in range(2, min(5, ws.max_row + 1)):
                print(f"\n  Zeile {row}:")
                for col_idx, col in enumerate(range(1, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        print(f"    {headers[col_idx]}: {cell_value}")

        print("-" * 80)

def delete_all_test_questions():
    """Löscht alle aktuellen Testfragen aus der Datenbank"""
    with app.app_context():
        count = TestQuestion.query.count()
        print(f"\nAktuelle Anzahl Testfragen in DB: {count}")

        if count > 0:
            print("Lösche alle Testfragen...")
            TestQuestion.query.delete()
            db.session.commit()
            print("✓ Alle Testfragen gelöscht!")
        else:
            print("Keine Testfragen zum Löschen vorhanden.")

def import_test_questions():
    """Importiert Testfragen aus dem Excel-File"""
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              'Testfragen_Weichenheizung_Komplett.xlsx')

    print(f"\nImportiere Testfragen aus: {excel_path}")
    print("=" * 80)

    wb = load_workbook(excel_path, data_only=True)

    with app.app_context():
        imported_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nVerarbeite Sheet: {sheet_name}")

            # Header-Zeile lesen
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers[col] = header.strip()

            print(f"Gefundene Spalten: {list(headers.values())}")

            # Spalten-Mapping basierend auf tatsächlichen Excel-Spaltennamen
            col_mapping = {}
            for col, header in headers.items():
                header_lower = header.lower()
                if header_lower == 'komponente':
                    col_mapping['komponente_typ'] = col
                elif header_lower == 'reihenfolge':
                    col_mapping['frage_nummer'] = col
                elif 'frage' in header_lower and 'text' in header_lower:
                    col_mapping['frage_text'] = col
                elif 'test' in header_lower and 'information' in header_lower:
                    col_mapping['test_information'] = col
                elif 'preset' in header_lower and 'lss' in header_lower and 'ch' in header_lower:
                    col_mapping['lss_ch'] = col
                elif 'preset' in header_lower and 'wh' in header_lower and 'lts' in header_lower:
                    col_mapping['wh_lts'] = col

            print(f"Spalten-Mapping: {col_mapping}")

            # Datenzeilen verarbeiten
            global_frage_nummer = 1  # Global unique Fragenummer

            for row in range(2, ws.max_row + 1):
                # Pflichtfelder auslesen
                komponente_typ = ws.cell(row=row, column=col_mapping.get('komponente_typ', 1)).value
                reihenfolge = ws.cell(row=row, column=col_mapping.get('frage_nummer', 2)).value  # "Reihenfolge" Spalte
                frage_text = ws.cell(row=row, column=col_mapping.get('frage_text', 3)).value

                # Leere Zeilen überspringen
                if not frage_text or not komponente_typ:
                    continue

                # Optionale Felder
                test_information = ws.cell(row=row, column=col_mapping.get('test_information', 4)).value
                wh_lts = ws.cell(row=row, column=col_mapping.get('wh_lts', 5)).value
                lss_ch = ws.cell(row=row, column=col_mapping.get('lss_ch', 6)).value

                # Testszenario aus Komponententyp generieren (da nicht im Excel vorhanden)
                testszenario = f"Standard-Tests {komponente_typ}"

                # Reihenfolge als Integer konvertieren
                try:
                    reihenfolge_int = int(reihenfolge) if reihenfolge else global_frage_nummer
                except (ValueError, TypeError):
                    reihenfolge_int = global_frage_nummer

                # Preset-Antworten erstellen (nur wenn nicht None und nicht leer)
                preset_antworten = {}
                if lss_ch and str(lss_ch).strip():
                    preset_antworten['lss_ch'] = str(lss_ch).strip()
                if wh_lts and str(wh_lts).strip():
                    preset_antworten['wh_lts'] = str(wh_lts).strip()

                # TestQuestion-Objekt erstellen
                test_question = TestQuestion(
                    komponente_typ=str(komponente_typ).strip(),
                    testszenario=testszenario,
                    frage_nummer=global_frage_nummer,  # Global unique
                    frage_text=str(frage_text).strip(),
                    test_information=str(test_information).strip() if test_information else None,
                    reihenfolge=reihenfolge_int,  # Aus Excel "Reihenfolge"-Spalte
                    preset_antworten=preset_antworten if preset_antworten else None
                )

                db.session.add(test_question)
                imported_count += 1

                print(f"  Importiere Frage {global_frage_nummer} (Komp: {komponente_typ}, Reih: {reihenfolge_int}): {frage_text[:40]}...")

                global_frage_nummer += 1

        # Alle Änderungen speichern
        db.session.commit()
        print(f"\n✓ Import abgeschlossen! {imported_count} Testfragen importiert.")

        # Verifizierung
        total_count = TestQuestion.query.count()
        print(f"✓ Verifizierung: {total_count} Testfragen in der Datenbank.")

if __name__ == '__main__':
    print("Testfragen Import-Script")
    print("=" * 80)

    # 1. Excel analysieren
    print("\n1. ANALYSE DES EXCEL-FILES")
    analyze_excel()

    # 2. Bestätigung einholen (wird vom Benutzer über Claude Code gemacht)
    print("\n\n2. DATENBANK-OPERATIONEN")
    print("=" * 80)

    # 3. Alte Daten löschen
    delete_all_test_questions()

    # 4. Neue Daten importieren
    import_test_questions()

    print("\n" + "=" * 80)
    print("✓ FERTIG!")
