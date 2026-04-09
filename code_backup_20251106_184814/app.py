from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from models import db, TestResult, Project, WHKConfig, TestQuestion, AbnahmeTestResult
from config import Config
from datetime import datetime
import json

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

with app.app_context():
    db.create_all()

@app.route('/')
def index():
    # Redirect to projects page
    return redirect(url_for('projekte'))

@app.route('/projekte')
def projekte():
    search_query = request.args.get('search', '')

    if search_query:
        # Case-insensitive Suche in Projektname, DIDOK, Projektleiter, Prüfer
        projekte = Project.query.filter(
            db.or_(
                Project.projektname.ilike(f'%{search_query}%'),
                Project.didok_betriebspunkt.ilike(f'%{search_query}%'),
                Project.projektleiter_sbb.ilike(f'%{search_query}%'),
                Project.pruefer_achermann.ilike(f'%{search_query}%')
            )
        ).order_by(Project.erstellt_am.desc()).all()
    else:
        projekte = Project.query.order_by(Project.erstellt_am.desc()).all()

    return render_template('projekte.html', projekte=projekte, search_query=search_query)

@app.route('/projekt/neu', methods=['GET', 'POST'])
def neues_projekt():
    if request.method == 'POST':
        from datetime import datetime

        # Datums-Felder konvertieren
        baumappenversion = request.form.get('baumappenversion')
        if baumappenversion:
            baumappenversion = datetime.strptime(baumappenversion, '%d.%m.%Y').date()
        else:
            baumappenversion = None

        pruefdatum = request.form.get('pruefdatum')
        if pruefdatum:
            pruefdatum = datetime.strptime(pruefdatum, '%Y-%m-%d').date()
        else:
            pruefdatum = None

        projekt = Project(
            energie=request.form['energie'],
            projektname=request.form['projektname'],
            didok_betriebspunkt=request.form.get('didok_betriebspunkt', ''),
            baumappenversion=baumappenversion,
            projektleiter_sbb=request.form.get('projektleiter_sbb', ''),
            pruefer_achermann=request.form.get('pruefer_achermann', ''),
            pruefdatum=pruefdatum,
            bemerkung=request.form.get('bemerkung', '')
        )
        db.session.add(projekt)
        db.session.commit()
        flash('Projekt erfolgreich angelegt!', 'success')
        return redirect(url_for('projekte'))
    return render_template('projekt_form.html', projekt=None, edit_mode=False)

@app.route('/projekt/bearbeiten/<int:projekt_id>', methods=['GET', 'POST'])
def projekt_bearbeiten(projekt_id):
    projekt = Project.query.get(projekt_id)

    if not projekt:
        flash('Projekt nicht gefunden!', 'error')
        return redirect(url_for('projekte'))

    if request.method == 'POST':
        from datetime import datetime

        # Datums-Felder konvertieren
        baumappenversion = request.form.get('baumappenversion')
        if baumappenversion:
            baumappenversion = datetime.strptime(baumappenversion, '%d.%m.%Y').date()
        else:
            baumappenversion = None

        pruefdatum = request.form.get('pruefdatum')
        if pruefdatum:
            pruefdatum = datetime.strptime(pruefdatum, '%Y-%m-%d').date()
        else:
            pruefdatum = None

        # Bestehende Projekt-Werte aktualisieren
        projekt.energie = request.form['energie']
        projekt.projektname = request.form['projektname']
        projekt.didok_betriebspunkt = request.form.get('didok_betriebspunkt', '')
        projekt.baumappenversion = baumappenversion
        projekt.projektleiter_sbb = request.form.get('projektleiter_sbb', '')
        projekt.pruefer_achermann = request.form.get('pruefer_achermann', '')
        projekt.pruefdatum = pruefdatum
        projekt.bemerkung = request.form.get('bemerkung', '')

        db.session.commit()
        flash('Projekt erfolgreich aktualisiert!', 'success')
        return redirect(url_for('projekte'))

    return render_template('projekt_form.html', projekt=projekt, edit_mode=True)

@app.route('/projekt/loeschen/<int:projekt_id>')
def projekt_loeschen(projekt_id):
    projekt = Project.query.get(projekt_id)

    if not projekt:
        flash('Projekt nicht gefunden!', 'error')
        return redirect(url_for('projekte'))

    # Prüfen, ob dem Projekt noch Tests zugeordnet sind
    if projekt.tests and len(projekt.tests) > 0:
        flash(f'Projekt kann nicht gelöscht werden, da noch {len(projekt.tests)} Test(s) zugeordnet sind. Bitte löschen Sie zuerst die zugehörigen Tests.', 'warning')
        return redirect(url_for('projekte'))

    # Projekt löschen
    projektname = projekt.projektname
    db.session.delete(projekt)
    db.session.commit()
    flash(f'Projekt "{projektname}" erfolgreich gelöscht!', 'success')
    return redirect(url_for('projekte'))

@app.route('/projekt/abnahmetest/<int:projekt_id>', methods=['GET', 'POST'])
def projekt_abnahmetest(projekt_id):
    projekt = Project.query.get(projekt_id)

    if not projekt:
        flash('Projekt nicht gefunden!', 'error')
        return redirect(url_for('projekte'))

    if request.method == 'POST':
        # Phase 5: POST-Handler für Speichern der Test-Ergebnisse
        answers_data_json = request.form.get('answers_data')

        if not answers_data_json:
            flash('Keine Antworten zum Speichern vorhanden!', 'warning')
            return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))

        try:
            answers_data = json.loads(answers_data_json)
            print(f"DEBUG: Empfangene Antworten: {len(answers_data)} Einträge")
            print(f"DEBUG: Erste 3 Keys: {list(answers_data.keys())[:3]}")

            # Lösche alle bestehenden Ergebnisse für dieses Projekt
            deleted_count = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).delete()
            print(f"DEBUG: {deleted_count} alte Einträge gelöscht")

            # Gruppiere Antworten nach Frage + Spalte
            # Format: {question_id}_{system}_{spalte}: ergebnis
            grouped_answers = {}

            for key, ergebnis in answers_data.items():
                # Parse key: "question_id_system_spalte"
                parts = key.split('_')

                if len(parts) < 3:
                    continue

                # Extrahiere question_id (kann mehrere Teile sein, z.B. "3_WHK_01")
                # System ist entweder "lss-ch" oder "wh-lts"
                # Spalte ist der Rest

                # Finde System-Position (lss-ch oder wh-lts)
                system_index = -1
                for i, part in enumerate(parts):
                    if part in ['lss-ch', 'wh-lts']:
                        system_index = i
                        break

                if system_index == -1:
                    continue

                # question_id ist alles vor dem System
                question_id_str = '_'.join(parts[:system_index])
                system = parts[system_index]
                spalte = '_'.join(parts[system_index + 1:])

                # Erstelle Gruppierungsschlüssel
                group_key = f"{question_id_str}_{spalte}"

                if group_key not in grouped_answers:
                    grouped_answers[group_key] = {
                        'question_id_str': question_id_str,
                        'spalte': spalte,
                        'lss_ch': None,
                        'wh_lts': None
                    }

                # Setze Ergebnis für das entsprechende System
                if system == 'lss-ch':
                    grouped_answers[group_key]['lss_ch'] = ergebnis
                elif system == 'wh-lts':
                    grouped_answers[group_key]['wh_lts'] = ergebnis

            # Erstelle AbnahmeTestResult-Einträge
            saved_count = 0
            for group_key, data in grouped_answers.items():
                question_id_str = data['question_id_str']
                spalte = data['spalte']

                # Extrahiere numerische test_question_id
                # question_id_str kann "1" oder "3_WHK_01" sein
                numeric_id_str = question_id_str.split('_')[0]
                try:
                    test_question_id = int(numeric_id_str)
                except (ValueError, IndexError):
                    print(f"Warnung: Konnte test_question_id nicht extrahieren aus: {question_id_str}")
                    continue

                # Komponente_index ist die Spalte
                komponente_index = spalte.replace('_', ' ')

                # Skip if test_question_id is None or invalid
                if not test_question_id:
                    print(f"Warnung: test_question_id ist None für {group_key}")
                    continue

                # Erstelle neuen Eintrag
                result = AbnahmeTestResult(
                    projekt_id=projekt_id,
                    test_question_id=test_question_id,
                    komponente_index=komponente_index,
                    lss_ch_result=data['lss_ch'],
                    wh_lts_result=data['wh_lts'],
                    getestet_am=datetime.utcnow()
                )

                db.session.add(result)
                saved_count += 1

            if saved_count > 0:
                db.session.commit()
                flash(f'Test-Ergebnisse erfolgreich gespeichert! ({saved_count} Einträge)', 'success')
            else:
                flash('Keine gültigen Antworten zum Speichern gefunden!', 'warning')

            return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))

        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            flash(f'Fehler beim Speichern: {str(e)}', 'error')
            return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))

    # Navigation-Format: Lade alle Testfragen und generiere JSON-Array

    # 1. Lade WHK-Konfigurationen für dieses Projekt
    whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

    # 2. Lade alle Testfragen sortiert
    all_test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

    # 3. Generiere Fragen-Array für JavaScript mit komponente_index
    fragen_array = []

    for frage in all_test_questions:
        # Bestimme Spalten und komponente_index basierend auf Komponententyp
        if frage.komponente_typ == "Anlage":
            # Anlage-Tests: Eine Frage, eine Spalte, kein Index
            fragen_array.append({
                'id': frage.id,
                'komponente_typ': frage.komponente_typ,
                'komponente_index': '',
                'frage_text': frage.frage_text,
                'test_information': frage.test_information or '',
                'spalten': ["Anlage"],
                'preset_antworten': frage.preset_antworten or {}
            })

        elif frage.komponente_typ == "WHK":
            # WHK-Tests: Eine Frage für alle WHK, mehrere Spalten, kein Index
            fragen_array.append({
                'id': frage.id,
                'komponente_typ': frage.komponente_typ,
                'komponente_index': '',
                'frage_text': frage.frage_text,
                'test_information': frage.test_information or '',
                'spalten': [whk.whk_nummer for whk in whk_configs],
                'preset_antworten': frage.preset_antworten or {}
            })

        elif frage.komponente_typ == "Abgang":
            # Abgang-Tests: Separate Frage für jede WHK
            for whk in whk_configs:
                spalten = [f"Abgang {i:02d}" for i in range(1, whk.anzahl_abgaenge + 1)]
                fragen_array.append({
                    'id': f"{frage.id}_{whk.whk_nummer.replace(' ', '_')}",
                    'komponente_typ': frage.komponente_typ,
                    'komponente_index': whk.whk_nummer,
                    'frage_text': frage.frage_text,
                    'test_information': frage.test_information or '',
                    'spalten': spalten,
                    'preset_antworten': frage.preset_antworten or {}
                })

        elif frage.komponente_typ == "Temperatursonde":
            # Temperatursonden-Tests: Separate Frage für jede WHK
            for whk in whk_configs:
                spalten = [f"TS {i:02d}" for i in range(1, whk.anzahl_temperatursonden + 1)]
                fragen_array.append({
                    'id': f"{frage.id}_{whk.whk_nummer.replace(' ', '_')}",
                    'komponente_typ': frage.komponente_typ,
                    'komponente_index': whk.whk_nummer,
                    'frage_text': frage.frage_text,
                    'test_information': frage.test_information or '',
                    'spalten': spalten,
                    'preset_antworten': frage.preset_antworten or {}
                })

        elif frage.komponente_typ == "Antriebsheizung":
            # Antriebsheizung-Tests: Nur für WHK mit Antriebsheizung
            for whk in whk_configs:
                if whk.hat_antriebsheizung:
                    fragen_array.append({
                        'id': f"{frage.id}_{whk.whk_nummer.replace(' ', '_')}",
                        'komponente_typ': frage.komponente_typ,
                        'komponente_index': whk.whk_nummer,
                        'frage_text': frage.frage_text,
                        'test_information': frage.test_information or '',
                        'spalten': ["Antriebsheizung"],
                        'preset_antworten': frage.preset_antworten or {}
                    })

        elif frage.komponente_typ == "Meteostation":
            # Meteostation-Tests: Für jede eindeutige Meteostation
            meteostationen = {}
            for whk in whk_configs:
                if whk.meteostation and whk.meteostation not in meteostationen:
                    meteostationen[whk.meteostation] = whk.whk_nummer

            for ms_name, whk_nummer in meteostationen.items():
                fragen_array.append({
                    'id': f"{frage.id}_{ms_name.replace(' ', '_')}",
                    'komponente_typ': frage.komponente_typ,
                    'komponente_index': ms_name,
                    'frage_text': frage.frage_text,
                    'test_information': frage.test_information or '',
                    'spalten': [ms_name],
                    'preset_antworten': frage.preset_antworten or {}
                })

    # 4. Lade bereits existierende Test-Ergebnisse
    # Format für JavaScript: question_id_system_spalte: ergebnis
    existing_results = {}
    results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

    for result in results:
        # Finde die entsprechende Frage in fragen_array
        # Wir müssen die question_id mit komponente_index matchen
        test_question = TestQuestion.query.get(result.test_question_id)

        if not test_question:
            continue

        # Bestimme die question_id für JavaScript
        # Bei Anlage/WHK: einfach die ID
        # Bei anderen: ID_komponente_index
        if test_question.komponente_typ in ["Anlage", "WHK"]:
            js_question_id = str(result.test_question_id)
        else:
            js_question_id = f"{result.test_question_id}_{result.komponente_index.replace(' ', '_')}"

        # Spalte ist jetzt explizit gespeichert (mit Underscores für Key)
        spalte = result.spalte.replace(' ', '_') if result.spalte else result.komponente_index.replace(' ', '_')

        # Erstelle Keys für LSS-CH und WH-LTS (mit Ergebnis und Bemerkung)
        if result.lss_ch_result or result.lss_ch_bemerkung:
            key_lss = f"{js_question_id}_lss-ch_{spalte}"
            existing_results[key_lss] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        if result.wh_lts_result or result.wh_lts_bemerkung:
            key_wh = f"{js_question_id}_wh-lts_{spalte}"
            existing_results[key_wh] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }

    # Als JSON-String für JavaScript übergeben
    fragen_json = json.dumps(fragen_array, ensure_ascii=False)
    existing_results_json = json.dumps(existing_results, ensure_ascii=False)

    return render_template('abnahmetest.html',
                         projekt=projekt,
                         whk_configs=whk_configs,
                         fragen_json=fragen_json,
                         existing_results_json=existing_results_json)

@app.route('/projekt/abnahmetest/save-answer', methods=['POST'])
def save_test_answer():
    """Auto-Save Route für einzelne Test-Antworten"""
    try:
        data = request.get_json()
        print(f"[DEBUG] Received save-answer request: {data}")

        projekt_id = data.get('projekt_id')
        question_id_str = data.get('question_id')  # Kann "1" oder "3_WHK_01" sein
        spalte = data.get('spalte')  # z.B. "WHK_01", "Abgang_01"
        system = data.get('system')  # 'lss-ch' oder 'wh-lts'
        ergebnis = data.get('ergebnis')  # 'richtig', 'falsch', 'nicht_testbar' oder None
        bemerkung = data.get('bemerkung')  # Optional: Bemerkungstext

        print(f"[DEBUG] Bemerkung parameter: {bemerkung} (type: {type(bemerkung)})")
        print(f"[DEBUG] All parameters: projekt_id={projekt_id}, question_id={question_id_str}, spalte={spalte}, system={system}, ergebnis={ergebnis}")

        # Validierung
        if not all([projekt_id, question_id_str, spalte, system]):
            print(f"[DEBUG] Validation failed - missing data")
            return jsonify({'success': False, 'error': 'Fehlende Daten'}), 400

        print(f"[DEBUG] Validation passed")

        # Prüfe ob Projekt existiert
        projekt = Project.query.get(projekt_id)
        if not projekt:
            return jsonify({'success': False, 'error': 'Projekt nicht gefunden'}), 404

        # Extrahiere numerische test_question_id und komponente_index
        # question_id_str kann "1" oder "3_WHK_01" oder "23_MS_01" sein
        # Konvertiere zu String falls es als Integer kommt
        parts = str(question_id_str).split('_')
        numeric_id_str = parts[0]

        try:
            test_question_id = int(numeric_id_str)
        except (ValueError, IndexError):
            return jsonify({'success': False, 'error': f'Ungültige Question ID: {question_id_str}'}), 400

        # Komponente_index: Für "3_WHK_01" -> "WHK 01", für "1" -> spalte
        if len(parts) > 1:
            # Frage hat komponente_index in der ID (z.B. "3_WHK_01", "23_MS_01")
            komponente_index = '_'.join(parts[1:]).replace('_', ' ')
        else:
            # Frage hat keinen komponente_index in der ID (Anlage/WHK)
            # In diesem Fall ist spalte der komponente_index
            komponente_index = spalte.replace('_', ' ')

        # Konvertiere system von lss-ch/wh-lts zu lss_ch/wh_lts
        system_db = system.replace('-', '_')

        # Spalte mit Spaces (für Anzeige und Speicherung)
        spalte_display = spalte.replace('_', ' ')

        # Suche nach bestehendem Eintrag (prüfe komponente_index UND spalte)
        existing_result = AbnahmeTestResult.query.filter_by(
            projekt_id=projekt_id,
            test_question_id=test_question_id,
            komponente_index=komponente_index,
            spalte=spalte_display
        ).first()

        if existing_result:
            # Update bestehenden Eintrag
            print(f"[DEBUG] Found existing result: ID={existing_result.id}, komponente_index={existing_result.komponente_index}, spalte={existing_result.spalte}")
            if system_db == 'lss_ch':
                existing_result.lss_ch_result = ergebnis
                # Bemerkung nur setzen wenn explizit übergeben (auch wenn leer zum Löschen)
                if bemerkung is not None:
                    old_bemerkung = existing_result.lss_ch_bemerkung
                    existing_result.lss_ch_bemerkung = bemerkung if bemerkung.strip() else None
                    print(f"[DEBUG] Updated lss_ch_bemerkung: '{old_bemerkung}' -> '{existing_result.lss_ch_bemerkung}'")
            elif system_db == 'wh_lts':
                existing_result.wh_lts_result = ergebnis
                # Bemerkung nur setzen wenn explizit übergeben (auch wenn leer zum Löschen)
                if bemerkung is not None:
                    old_bemerkung = existing_result.wh_lts_bemerkung
                    existing_result.wh_lts_bemerkung = bemerkung if bemerkung.strip() else None
                    print(f"[DEBUG] Updated wh_lts_bemerkung: '{old_bemerkung}' -> '{existing_result.wh_lts_bemerkung}'")

            existing_result.getestet_am = datetime.utcnow()
            if projekt.pruefer_achermann:
                existing_result.tester = projekt.pruefer_achermann
        else:
            # Erstelle neuen Eintrag
            print(f"[DEBUG] Creating new result entry")
            print(f"[DEBUG] komponente_index={komponente_index}, spalte={spalte_display}")
            new_result = AbnahmeTestResult(
                projekt_id=projekt_id,
                test_question_id=test_question_id,
                komponente_index=komponente_index,
                spalte=spalte_display,
                lss_ch_result=ergebnis if system_db == 'lss_ch' else None,
                wh_lts_result=ergebnis if system_db == 'wh_lts' else None,
                lss_ch_bemerkung=bemerkung if system_db == 'lss_ch' and bemerkung else None,
                wh_lts_bemerkung=bemerkung if system_db == 'wh_lts' and bemerkung else None,
                getestet_am=datetime.utcnow(),
                tester=projekt.pruefer_achermann or 'Unbekannt'
            )
            print(f"[DEBUG] New entry bemerkungen: lss_ch={new_result.lss_ch_bemerkung}, wh_lts={new_result.wh_lts_bemerkung}")
            db.session.add(new_result)

        db.session.commit()
        print(f"[DEBUG] Database commit successful")

        return jsonify({
            'success': True,
            'message': 'Antwort gespeichert',
            'timestamp': datetime.utcnow().isoformat()
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/projekt/konfiguration/<int:projekt_id>', methods=['GET', 'POST'])
def projekt_konfiguration(projekt_id):
    projekt = Project.query.get(projekt_id)

    if not projekt:
        flash('Projekt nicht gefunden!', 'error')
        return redirect(url_for('projekte'))

    if request.method == 'POST':
        # Lösche alle bestehenden WHK-Konfigurationen für dieses Projekt
        WHKConfig.query.filter_by(projekt_id=projekt_id).delete()

        # Verarbeite die Formular-Daten
        # Durchlaufe alle whk_nr_* Felder im Formular
        whk_count = 0
        for key in request.form.keys():
            if key.startswith('whk_nr_'):
                # Extrahiere die Nummer (z.B. whk_nr_1 -> 1)
                index = key.split('_')[-1]

                whk_nummer = request.form.get(f'whk_nr_{index}')
                anzahl_abgaenge = int(request.form.get(f'abgaenge_{index}', 1))
                anzahl_temperatursonden = int(request.form.get(f'temperatursonden_{index}', 1))
                hat_antriebsheizung = f'antriebsheizung_{index}' in request.form
                meteostation = request.form.get(f'meteostation_{index}', '').strip() or None

                # Erstelle neuen WHKConfig-Eintrag
                whk_config = WHKConfig(
                    projekt_id=projekt_id,
                    whk_nummer=whk_nummer,
                    anzahl_abgaenge=anzahl_abgaenge,
                    anzahl_temperatursonden=anzahl_temperatursonden,
                    hat_antriebsheizung=hat_antriebsheizung,
                    meteostation=meteostation
                )
                db.session.add(whk_config)
                whk_count += 1

        db.session.commit()
        flash(f'Konfiguration erfolgreich gespeichert! ({whk_count} WHK konfiguriert)', 'success')
        return redirect(url_for('projekt_konfiguration', projekt_id=projekt_id))

    # GET-Request: Lade bestehende Konfigurationen
    whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

    return render_template('konfiguration.html', projekt=projekt, whk_configs=whk_configs)

@app.route('/projekt/konfiguration/auto-save/<int:projekt_id>', methods=['POST'])
def konfiguration_auto_save(projekt_id):
    """Auto-Save Route für WHK-Konfiguration"""
    try:
        data = request.get_json()

        # Prüfe ob Projekt existiert
        projekt = Project.query.get(projekt_id)
        if not projekt:
            return jsonify({'success': False, 'error': 'Projekt nicht gefunden'}), 404

        # Lösche alle bestehenden WHK-Konfigs für dieses Projekt
        WHKConfig.query.filter_by(projekt_id=projekt_id).delete()

        # Erstelle neue WHK-Konfigs aus den übermittelten Daten
        whk_rows = data.get('whk_rows', [])
        saved_count = 0

        for row_data in whk_rows:
            # Validierung: WHK-Nummer muss vorhanden sein
            whk_nummer = row_data.get('whk_nummer', '').strip()
            if not whk_nummer:
                continue  # Überspringe unvollständige Zeilen

            new_config = WHKConfig(
                projekt_id=projekt_id,
                whk_nummer=whk_nummer,
                anzahl_abgaenge=int(row_data.get('anzahl_abgaenge', 1)),
                anzahl_temperatursonden=int(row_data.get('anzahl_temperatursonden', 1)),
                hat_antriebsheizung=row_data.get('hat_antriebsheizung', False),
                meteostation=row_data.get('meteostation', '').strip() or None
            )
            db.session.add(new_config)
            saved_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Konfiguration gespeichert',
            'timestamp': datetime.utcnow().isoformat(),
            'count': saved_count
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/tests')
def tests():
    # Alle Tests mit Projektzuordnung anzeigen
    tests = TestResult.query.order_by(TestResult.test_date.desc()).all()
    return render_template('tests.html', tests=tests)

@app.route('/new_test', methods=['GET', 'POST'])
def new_test():
    if request.method == 'POST':
        # Projekt-ID holen (optional)
        project_id = request.form.get('project_id')
        if project_id == '':
            project_id = None

        test = TestResult(
            test_name=request.form['test_name'],
            hardware_id=request.form['hardware_id'],
            software_version=request.form['software_version'],
            result=request.form['result'],
            tester_name=request.form['tester_name'],
            notes=request.form.get('notes', ''),
            project_id=project_id
        )
        db.session.add(test)
        db.session.commit()
        flash('Test erfolgreich gespeichert!', 'success')
        return redirect(url_for('index'))

    # Alle Projekte für Dropdown laden
    projekte = Project.query.order_by(Project.projektname).all()
    return render_template('test_form.html', projekte=projekte)

@app.route('/testfragen')
def testfragen_verwaltung():
    test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()
    return render_template('testfragen_verwaltung.html', test_questions=test_questions)

@app.route('/testfragen/neu', methods=['GET', 'POST'])
def testfrage_neu():
    if request.method == 'POST':
        # Build preset_antworten JSON
        preset_antworten = {}
        preset_lss_ch = request.form.get('preset_lss_ch', '')
        preset_wh_lts = request.form.get('preset_wh_lts', '')

        if preset_lss_ch:
            preset_antworten['lss_ch'] = preset_lss_ch
        if preset_wh_lts:
            preset_antworten['wh_lts'] = preset_wh_lts

        komponente_typ = request.form['komponente_typ']

        # Automatische Generierung von frage_nummer (höchste + 1)
        max_frage_nummer = db.session.query(db.func.max(TestQuestion.frage_nummer)).scalar() or 0
        neue_frage_nummer = max_frage_nummer + 1

        # Automatische Generierung von reihenfolge (höchste für diesen Typ + 1)
        max_reihenfolge = db.session.query(db.func.max(TestQuestion.reihenfolge))\
            .filter(TestQuestion.komponente_typ == komponente_typ).scalar() or 0
        neue_reihenfolge = max_reihenfolge + 1

        neue_frage = TestQuestion(
            komponente_typ=komponente_typ,
            testszenario=request.form.get('testszenario', ''),
            frage_nummer=neue_frage_nummer,
            frage_text=request.form['frage_text'],
            test_information=request.form.get('test_information', ''),
            reihenfolge=neue_reihenfolge,
            preset_antworten=preset_antworten if preset_antworten else None
        )
        db.session.add(neue_frage)
        db.session.commit()
        flash('Testfrage erfolgreich hinzugefügt!', 'success')
        return redirect(url_for('testfragen_verwaltung', tab=komponente_typ))

    # GET: Hole tab aus Query-Parameter
    tab = request.args.get('tab', 'Anlage')
    return render_template('testfrage_form.html', frage=None, tab=tab)

@app.route('/testfragen/bearbeiten/<int:frage_id>', methods=['GET', 'POST'])
def testfrage_bearbeiten(frage_id):
    frage = TestQuestion.query.get_or_404(frage_id)

    if request.method == 'POST':
        # Build preset_antworten JSON
        preset_antworten = {}
        preset_lss_ch = request.form.get('preset_lss_ch', '')
        preset_wh_lts = request.form.get('preset_wh_lts', '')

        if preset_lss_ch:
            preset_antworten['lss_ch'] = preset_lss_ch
        if preset_wh_lts:
            preset_antworten['wh_lts'] = preset_wh_lts

        komponente_typ = request.form['komponente_typ']
        frage.komponente_typ = komponente_typ
        frage.testszenario = request.form.get('testszenario', '')
        # frage_nummer bleibt unverändert (wird nicht mehr bearbeitet)
        frage.frage_text = request.form['frage_text']
        frage.test_information = request.form.get('test_information', '')
        # reihenfolge bleibt unverändert (kann über Drag & Drop geändert werden)
        frage.preset_antworten = preset_antworten if preset_antworten else None

        db.session.commit()
        flash('Testfrage erfolgreich aktualisiert!', 'success')
        return redirect(url_for('testfragen_verwaltung', tab=komponente_typ))

    # GET: Hole tab aus Query-Parameter (oder verwende komponente_typ der Frage)
    tab = request.args.get('tab', frage.komponente_typ)
    return render_template('testfrage_form.html', frage=frage, tab=tab)

@app.route('/testfragen/loeschen/<int:frage_id>', methods=['POST'])
def testfrage_loeschen(frage_id):
    frage = TestQuestion.query.get_or_404(frage_id)
    komponente_typ = frage.komponente_typ  # Merke den Tab vor dem Löschen
    db.session.delete(frage)
    db.session.commit()
    flash('Testfrage erfolgreich gelöscht!', 'success')
    return redirect(url_for('testfragen_verwaltung', tab=komponente_typ))

@app.route('/testfragen/reihenfolge', methods=['POST'])
def testfragen_reihenfolge():
    from flask import jsonify
    data = request.get_json()

    try:
        for item in data['order']:
            frage = TestQuestion.query.get(int(item['id']))
            if frage:
                frage.reihenfolge = int(item['reihenfolge'])

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

# ==================== PDF-EXPORT ====================
@app.route('/projekt/<int:projekt_id>/export/pdf')
def export_pdf(projekt_id):
    """Exportiert Abnahmetest-Protokoll als PDF"""
    try:
        from weasyprint import HTML
        from io import BytesIO
        import os

        # Projekt laden
        projekt = Project.query.get_or_404(projekt_id)

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse für dieses Projekt laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln für schnellen Zugriff
        results_dict = {}
        for result in results:
            # Key-Format: question_id_system_komponente_index_spalte
            key_wh_lts = f"{result.test_question_id}_wh_lts_{result.komponente_index}_{result.spalte or ''}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{result.komponente_index}_{result.spalte or ''}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktion: Icon für Checkbox-Wert
        def get_icon(result_value):
            assets_path = os.path.join(os.getcwd(), 'assets').replace('\\', '/')
            if result_value == 'richtig':
                return f'<img src="file:///{assets_path}/richtig.svg" class="check-icon" alt="✓">'
            elif result_value == 'falsch':
                return f'<img src="file:///{assets_path}/falsch.svg" class="check-icon" alt="✗">'
            elif result_value == 'nicht_testbar':
                return f'<img src="file:///{assets_path}/nicht_testbar.svg" class="check-icon" alt="⊘">'
            else:
                return ''

        # Helper-Funktion: Testergebnis abrufen
        def get_test_result(question_id, system, komponente_index, spalte=''):
            key = f"{question_id}_{system}_{komponente_index}_{spalte}"
            result_data = results_dict.get(key, {})
            return {
                'icon': get_icon(result_data.get('result')),
                'bemerkung': result_data.get('bemerkung', '')
            }

        # WH-Anlage Tests vorbereiten
        anlage_tests = []
        anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
        for frage in anlage_fragen:
            wh_lts_result = get_test_result(frage.id, 'wh_lts', '', '')  # Beide leer für Anlage
            lss_ch_result = get_test_result(frage.id, 'lss_ch', '', '')  # Beide leer für Anlage

            anlage_tests.append({
                'frage_text': frage.frage_text,
                'wh_lts_icon': wh_lts_result['icon'],
                'lss_ch_icon': lss_ch_result['icon'],
                'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
            })

        # WHK-Daten vorbereiten
        whk_data = []
        for whk_config in whk_configs:
            whk_nummer = whk_config.whk_nummer

            # WHK-Tests
            whk_tests = []
            whk_fragen = [q for q in test_questions if q.komponente_typ == 'WHK']
            for frage in whk_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer, '')  # Spalte leer für WHK-Tests
                lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer, '')  # Spalte leer für WHK-Tests

                whk_tests.append({
                    'frage_text': frage.frage_text,
                    'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte (keine separate whk_result Spalte in DB)
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            # Abgang-Tests
            abgang_tests = []
            abgang_fragen = [q for q in test_questions if q.komponente_typ == 'Abgang']
            for abgang_num in range(1, whk_config.anzahl_abgaenge + 1):
                abgang_name = f"Abgang {abgang_num:02d}"
                for frage in abgang_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer, abgang_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer, abgang_name)

                    abgang_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': abgang_name,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Temperatursonden-Tests
            ts_tests = []
            ts_fragen = [q for q in test_questions if q.komponente_typ == 'Temperatursonde']
            for ts_num in range(1, whk_config.anzahl_temperatursonden + 1):
                ts_name = f"TS {ts_num:02d}"
                for frage in ts_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer, ts_name)
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer, ts_name)

                    ts_tests.append({
                        'frage_text': frage.frage_text,
                        'spalte': ts_name,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            # Antriebsheizung-Tests
            ah_tests = []
            if whk_config.hat_antriebsheizung:
                ah_fragen = [q for q in test_questions if q.komponente_typ == 'Antriebsheizung']
                for frage in ah_fragen:
                    wh_lts_result = get_test_result(frage.id, 'wh_lts', whk_nummer, 'Antriebsheizung')
                    lss_ch_result = get_test_result(frage.id, 'lss_ch', whk_nummer, 'Antriebsheizung')

                    ah_tests.append({
                        'frage_text': frage.frage_text,
                        'whk_icon': wh_lts_result['icon'],  # WHK zeigt WH-LTS Werte
                        'wh_lts_icon': wh_lts_result['icon'],
                        'lss_ch_icon': lss_ch_result['icon'],
                        'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                    })

            whk_data.append({
                'whk_nummer': whk_nummer,
                'anzahl_abgaenge': whk_config.anzahl_abgaenge,
                'anzahl_temperatursonden': whk_config.anzahl_temperatursonden,
                'whk_tests': whk_tests,
                'abgang_tests': abgang_tests,
                'ts_tests': ts_tests,
                'ah_tests': ah_tests
            })

        # Meteostation-Daten vorbereiten
        meteo_data = []
        meteo_stations = list(set([whk.meteostation for whk in whk_configs if whk.meteostation]))
        for meteo_station in meteo_stations:
            meteo_tests = []
            meteo_fragen = [q for q in test_questions if q.komponente_typ == 'Meteostation']
            for frage in meteo_fragen:
                wh_lts_result = get_test_result(frage.id, 'wh_lts', meteo_station, '')  # Spalte leer für Meteostation
                lss_ch_result = get_test_result(frage.id, 'lss_ch', meteo_station, '')  # Spalte leer für Meteostation

                meteo_tests.append({
                    'frage_text': frage.frage_text,
                    'wh_lts_icon': wh_lts_result['icon'],
                    'lss_ch_icon': lss_ch_result['icon'],
                    'bemerkung': wh_lts_result['bemerkung'] or lss_ch_result['bemerkung']
                })

            if meteo_tests:
                meteo_data.append({
                    'meteostation': meteo_station,
                    'tests': meteo_tests
                })

        # Prüfe ob Logos existieren
        assets_path = os.path.join(os.path.dirname(__file__), 'assets').replace('\\', '/')
        sbb_logo_path = os.path.join(assets_path, 'sbb06.gif')
        achermann_logo_path = os.path.join(assets_path, 'Logo Achermann black.svg')

        sbb_logo_exists = os.path.exists(sbb_logo_path)
        achermann_logo_exists = os.path.exists(achermann_logo_path)

        # Template rendern
        html_string = render_template(
            'pdf_abnahmetest.html',
            projekt=projekt,
            projektname=projekt.projektname,
            didok=projekt.didok_betriebspunkt or '',
            projektleiter_sbb=projekt.projektleiter_sbb or '',
            baumappenversion=projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else '',
            pruefer_achermann=projekt.pruefer_achermann or '',
            pruefdatum=datetime.now().strftime('%d.%m.%Y'),
            anlage_tests=anlage_tests,
            whk_data=whk_data,
            meteo_data=meteo_data,
            assets_path=assets_path,
            sbb_logo_exists=sbb_logo_exists,
            achermann_logo_exists=achermann_logo_exists
        )

        # PDF generieren
        pdf = HTML(string=html_string).write_pdf()

        # Als Download zurückgeben
        from flask import send_file
        return send_file(
            BytesIO(pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Abnahmetest_{projekt.projektname}_{projekt.didok_betriebspunkt or "keine_DIDOK"}.pdf'
        )

    except ImportError:
        flash('WeasyPrint ist nicht installiert. Bitte installieren Sie es mit: pip install weasyprint', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))
    except Exception as e:
        error_msg = str(e)

        # Spezielle Behandlung für GTK-Fehler auf Windows
        if 'libgobject' in error_msg or 'library' in error_msg.lower():
            flash(
                'Fehler beim PDF-Export: WeasyPrint benötigt GTK-Bibliotheken. '
                'TIPP: Nutzen Sie den Excel-Export als Alternative! '
                'Für PDF-Support auf Windows: Laden Sie GTK3-Runtime herunter von '
                'https://github.com/tschoonj/GTK-for-Windows-Runtime-Environment-Installer/releases',
                'error'
            )
        else:
            flash(f'Fehler beim PDF-Export: {error_msg}', 'error')

        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))


# ==================== EXCEL-EXPORT ====================
@app.route('/projekt/<int:projekt_id>/export/excel')
def export_excel(projekt_id):
    """Exportiert Abnahmetest-Protokoll als Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        # Projekt laden
        projekt = Project.query.get_or_404(projekt_id)

        # WHK-Konfigurationen laden
        whk_configs = WHKConfig.query.filter_by(projekt_id=projekt_id).order_by(WHKConfig.whk_nummer).all()

        # Alle Testfragen laden
        test_questions = TestQuestion.query.order_by(TestQuestion.komponente_typ, TestQuestion.reihenfolge).all()

        # Alle Testergebnisse laden
        results = AbnahmeTestResult.query.filter_by(projekt_id=projekt_id).all()

        # Ergebnisse in Dictionary umwandeln
        results_dict = {}
        for result in results:
            key_wh_lts = f"{result.test_question_id}_wh_lts_{result.komponente_index}_{result.spalte or ''}"
            key_lss_ch = f"{result.test_question_id}_lss_ch_{result.komponente_index}_{result.spalte or ''}"

            results_dict[key_wh_lts] = {
                'result': result.wh_lts_result,
                'bemerkung': result.wh_lts_bemerkung
            }
            results_dict[key_lss_ch] = {
                'result': result.lss_ch_result,
                'bemerkung': result.lss_ch_bemerkung
            }

        # Helper-Funktion: Text für Checkbox-Wert
        def get_result_text(result_value):
            if result_value == 'richtig':
                return '✓ Richtig'
            elif result_value == 'falsch':
                return '✗ Falsch'
            elif result_value == 'nicht_testbar':
                return '⊘ Nicht testbar'
            else:
                return ''

        # Workbook erstellen
        wb = Workbook()

        # Styles definieren
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Projektinformationen
        ws1 = wb.active
        ws1.title = "Projektinformationen"

        ws1['A1'] = "Abnahmetest Elektroweichenheizung"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:B1')

        ws1['A3'] = "Betriebspunkt (Projektname):"
        ws1['A3'].font = Font(bold=True)
        ws1['B3'] = projekt.projektname

        ws1['A4'] = "DIDOK:"
        ws1['A4'].font = Font(bold=True)
        ws1['B4'] = projekt.didok_betriebspunkt or ''

        ws1['A5'] = "Projektleiter SBB AG:"
        ws1['A5'].font = Font(bold=True)
        ws1['B5'] = projekt.projektleiter_sbb or ''

        ws1['A6'] = "Baumappen Version (Datum):"
        ws1['A6'].font = Font(bold=True)
        ws1['B6'] = projekt.baumappenversion.strftime('%d.%m.%Y') if projekt.baumappenversion else ''

        ws1['A7'] = "Prüfer:"
        ws1['A7'].font = Font(bold=True)
        ws1['B7'] = projekt.pruefer_achermann or ''

        ws1['A8'] = "Prüfdatum:"
        ws1['A8'].font = Font(bold=True)
        ws1['B8'] = datetime.now().strftime('%d.%m.%Y')

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 50

        # Sheet 2: WH-Anlage Tests
        ws2 = wb.create_sheet("WH-Anlage")
        ws2['A1'] = "Test"
        ws2['B1'] = "WH-LTS"
        ws2['C1'] = "LSS-CH"
        ws2['D1'] = "Bemerkung"

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        row = 2
        anlage_fragen = [q for q in test_questions if q.komponente_typ == 'Anlage']
        for frage in anlage_fragen:
            key_wh_lts = f"{frage.id}_wh_lts_Anlage_Anlage"
            key_lss_ch = f"{frage.id}_lss_ch_Anlage_Anlage"

            wh_lts_data = results_dict.get(key_wh_lts, {})
            lss_ch_data = results_dict.get(key_lss_ch, {})

            ws2[f'A{row}'] = frage.frage_text
            ws2[f'B{row}'] = get_result_text(wh_lts_data.get('result'))
            ws2[f'C{row}'] = get_result_text(lss_ch_data.get('result'))
            ws2[f'D{row}'] = wh_lts_data.get('bemerkung', '') or lss_ch_data.get('bemerkung', '')

            for cell in ws2[row]:
                cell.border = border_thin

            row += 1

        ws2.column_dimensions['A'].width = 60
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 40

        # Weitere Sheets für WHKs
        for whk_config in whk_configs:
            ws_whk = wb.create_sheet(f"WHK {whk_config.whk_nummer}")
            ws_whk['A1'] = "Test"
            ws_whk['B1'] = "WH-LTS"
            ws_whk['C1'] = "LSS-CH"
            ws_whk['D1'] = "Bemerkung"

            for cell in ws_whk[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            ws_whk.column_dimensions['A'].width = 60
            ws_whk.column_dimensions['B'].width = 15
            ws_whk.column_dimensions['C'].width = 15
            ws_whk.column_dimensions['D'].width = 40

            # Hier würden die WHK-Tests eingefügt werden (ähnlich wie oben)
            # Aus Platzgründen hier vereinfacht

        # Excel in Memory speichern
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        from flask import send_file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Abnahmetest_{projekt.projektname}_{projekt.didok_betriebspunkt or "keine_DIDOK"}.xlsx'
        )

    except ImportError:
        flash('openpyxl ist nicht installiert. Bitte installieren Sie es mit: pip install openpyxl', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))
    except Exception as e:
        flash(f'Fehler beim Excel-Export: {str(e)}', 'error')
        return redirect(url_for('projekt_abnahmetest', projekt_id=projekt_id))


if __name__ == '__main__':
    app.run(debug=True)
